#!/usr/bin/env python3
"""
ETL Intermediate to PostgreSQL - Sprint 1 PostgreSQL Migration
============================================================

Este script carga archivos intermedios validados a la base de datos PostgreSQL
con soporte PostGIS para coordenadas geoespaciales.

Flujo: Excel Validado → PostgreSQL + PostGIS

Author: Claude Code for Citrino
Date: October 2025
"""

import os
import re
import json
import logging
import psycopg2
import psycopg2.extras
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/postgres/logs/etl_intermediate_to_postgres.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class IntermediateToPostgresETL:
    """ETL para cargar archivos intermedios validados a PostgreSQL."""

    def __init__(self, validated_dir: str, logs_dir: str, backups_dir: str):
        """
        Inicializar el ETL.

        Args:
            validated_dir: Directorio con archivos validados
            logs_dir: Directorio para logs de transacciones
            backups_dir: Directorio para respaldos
        """
        self.validated_dir = Path(validated_dir)
        self.logs_dir = Path(logs_dir)
        self.backups_dir = Path(backups_dir)

        # Asegurar que los directorios existan
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        self.backups_dir.mkdir(parents=True, exist_ok=True)

        # Configuración de base de datos
        self.db_config = {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME', 'citrino'),
            'user': os.getenv('DB_USER', 'postgres'),
            'password': os.getenv('DB_PASSWORD', '')
        }

        # Estadísticas del proceso
        self.stats = {
            'agentes_cargados': 0,
            'propiedades_cargadas': 0,
            'servicios_cargados': 0,
            'errores_carga': 0,
            'transacciones_exitosas': 0,
            'archivos_procesados': 0
        }

        # Conexión a base de datos
        self.connection = None

    def conectar_db(self) -> bool:
        """
        Establecer conexión a la base de datos PostgreSQL.

        Returns:
            True si la conexión es exitosa
        """
        try:
            self.connection = psycopg2.connect(**self.db_config)
            self.connection.autocommit = False
            logger.info(f"Conectado a PostgreSQL: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")
            return True
        except Exception as e:
            logger.error(f"Error conectando a PostgreSQL: {e}")
            return False

    def desconectar_db(self):
        """Cerrar conexión a la base de datos."""
        if self.connection:
            self.connection.close()
            logger.info("Conexión a PostgreSQL cerrada")

    def ejecutar_query(self, query: str, params: Optional[Tuple] = None) -> Any:
        """
        Ejecutar query en la base de datos.

        Args:
            query: Query SQL a ejecutar
            params: Parámetros del query

        Returns:
            Resultado del query
        """
        try:
            with self.connection.cursor() as cursor:
                cursor.execute(query, params)
                if query.strip().upper().startswith('SELECT'):
                    return cursor.fetchall()
                else:
                    self.connection.commit()
                    return cursor.rowcount
        except Exception as e:
            self.connection.rollback()
            logger.error(f"Error ejecutando query: {e}")
            raise

    def crear_respaldo_tabla(self, nombre_tabla: str) -> str:
        """
        Crear respaldo de tabla antes de cargar datos.

        Args:
            nombre_tabla: Nombre de la tabla a respaldar

        Returns:
            Nombre del archivo de respaldo
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"backup_{nombre_tabla}_{timestamp}.sql"
        backup_path = self.backups_dir / backup_filename

        try:
            # Exportar tabla a SQL
            with open(backup_path, 'w', encoding='utf-8') as f:
                f.write(f"-- Respaldo de tabla {nombre_tabla}\n")
                f.write(f"-- Fecha: {datetime.now()}\n\n")

                # Exportar estructura
                with self.connection.cursor() as cursor:
                    cursor.execute(f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = '{nombre_tabla}'")
                    columns = cursor.fetchall()

                    cursor.execute(f"SELECT * FROM {nombre_tabla}")
                    rows = cursor.fetchall()

                    # Generar INSERT statements
                    for row in rows:
                        values = ', '.join([f"'{str(v).replace(\"'\", \"''\")}'" if v is not None else 'NULL' for v in row])
                        f.write(f"INSERT INTO {nombre_tabla} VALUES ({values});\n")

            logger.info(f"Respaldo creado: {backup_path}")
            return str(backup_path)

        except Exception as e:
            logger.error(f"Error creando respaldo de {nombre_tabla}: {e}")
            return ""

    def cargar_agentes(self, archivo_agentes: Path) -> Dict[str, Any]:
        """
        Cargar agentes consolidados a PostgreSQL.

        Args:
            archivo_agentes: Archivo Excel con agentes consolidados

        Returns:
            Resultado de la carga
        """
        logger.info(f"Cargando agentes desde: {archivo_agentes.name}")

        try:
            # Leer Excel
            df = pd.read_excel(archivo_agentes, sheet_name="Agentes Consolidados", engine='openpyxl')

            # Crear respaldo
            backup_file = self.crear_respaldo_tabla('agentes')

            # Preparar datos para inserción
            agentes_cargados = 0
            errores = []

            for idx, row in df.iterrows():
                try:
                    # Insertar agente
                    query = """
                    INSERT INTO agentes (nombre, telefono, email, empresa, fecha_registro)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (email) DO UPDATE SET
                        nombre = EXCLUDED.nombre,
                        telefono = EXCLUDED.telefono,
                        empresa = EXCLUDED.empresa
                    RETURNING id
                    """

                    params = (
                        row.get('nombre'),
                        row.get('telefono'),
                        row.get('email'),
                        row.get('empresa'),
                        datetime.now()
                    )

                    with self.connection.cursor() as cursor:
                        cursor.execute(query, params)
                        agente_id = cursor.fetchone()[0]
                        agentes_cargados += 1

                except Exception as e:
                    errores.append({
                        'fila': idx + 1,
                        'error': str(e),
                        'datos': row.to_dict()
                    })
                    logger.error(f"Error cargando agente fila {idx+1}: {e}")

            self.stats['agentes_cargados'] = agentes_cargados

            return {
                'cargados': agentes_cargados,
                'errores': len(errores),
                'detalle_errores': errores,
                'backup_file': backup_file
            }

        except Exception as e:
            logger.error(f"Error procesando archivo de agentes: {e}")
            return {'error': str(e), 'cargados': 0, 'errores': 1}

    def cargar_propiedades(self, archivo_propiedades: Path) -> Dict[str, Any]:
        """
        Cargar propiedades a PostgreSQL.

        Args:
            archivo_propiedades: Archivo Excel con propiedades

        Returns:
            Resultado de la carga
        """
        logger.info(f"Cargando propiedades desde: {archivo_propiedades.name}")

        try:
            # Leer Excel
            df = pd.read_excel(archivo_propiedades, sheet_name="Propiedades", engine='openpyxl')

            # Crear respaldo
            backup_file = self.crear_respaldo_tabla('propiedades')

            # Preparar datos para inserción
            propiedades_cargadas = 0
            errores = []

            for idx, row in df.iterrows():
                try:
                    # Preparar coordenadas PostGIS
                    coordenadas_postgis = None
                    if (not pd.isna(row.get('latitud')) and not pd.isna(row.get('longitud')) and
                        row.get('coordenadas_validas', False)):
                        coordenadas_postgis = f"ST_GeographyFromText('SRID=4326;POINT({row['longitud']} {row['latitud']})')"

                    # Buscar agente si existe información de contacto
                    agente_id = None
                    # Aquí se podría implementar lógica para buscar agente por nombre/email/telefono

                    # Insertar propiedad
                    if coordenadas_postgis:
                        query = f"""
                        INSERT INTO propiedades (
                            titulo, descripcion, tipo_propiedad, estado_propiedad,
                            precio_usd, precio_usd_m2, direccion, zona, uv, manzana, lote,
                            superficie_total, superficie_construida, num_dormitorios,
                            num_banos, num_garajes, coordenadas, fecha_publicacion,
                            fecha_scraping, ultima_actualizacion, proveedor_datos,
                            codigo_proveedor, url_origen, coordenadas_validas, datos_completos
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, {coordenadas_postgis}, %s, %s, %s, %s, %s, %s, %s, %s
                        )
                        """
                    else:
                        query = """
                        INSERT INTO propiedades (
                            titulo, descripcion, tipo_propiedad, estado_propiedad,
                            precio_usd, precio_usd_m2, direccion, zona, uv, manzana, lote,
                            superficie_total, superficie_construida, num_dormitorios,
                            num_banos, num_garajes, coordenadas, fecha_publicacion,
                            fecha_scraping, ultima_actualizacion, proveedor_datos,
                            codigo_proveedor, url_origen, coordenadas_validas, datos_completos
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                        )
                        """

                    params = (
                        row.get('titulo'),
                        row.get('descripcion'),
                        row.get('tipo_propiedad'),
                        row.get('estado_propiedad', 'En Venta'),
                        row.get('precio_usd'),
                        row.get('precio_usd_m2'),
                        row.get('direccion'),
                        row.get('zona'),
                        row.get('uv'),
                        row.get('manzana'),
                        row.get('lote'),
                        row.get('superficie_total'),
                        row.get('superficie_construida'),
                        row.get('num_dormitorios'),
                        row.get('num_banos'),
                        row.get('num_garajes'),
                        None,  # coordenadas se maneja aparte
                        row.get('fecha_publicacion'),
                        row.get('fecha_scraping', datetime.now()),
                        datetime.now(),
                        row.get('proveedor_datos', 'excel_validado'),
                        row.get('codigo_proveedor'),
                        row.get('url_origen'),
                        row.get('coordenadas_validas', False),
                        True  # datos_completos (viene de archivo validado)
                    )

                    with self.connection.cursor() as cursor:
                        if coordenadas_postgis:
                            cursor.execute(query, params)
                        else:
                            cursor.execute(query, params + (None,))
                        propiedades_cargadas += 1

                except Exception as e:
                    errores.append({
                        'fila': idx + 1,
                        'error': str(e),
                        'datos': {
                            'titulo': row.get('titulo'),
                            'precio': row.get('precio_usd'),
                            'zona': row.get('zona')
                        }
                    })
                    logger.error(f"Error cargando propiedad fila {idx+1}: {e}")

            self.stats['propiedades_cargadas'] = propiedades_cargadas

            return {
                'cargadas': propiedades_cargadas,
                'errores': len(errores),
                'detalle_errores': errores,
                'backup_file': backup_file
            }

        except Exception as e:
            logger.error(f"Error procesando archivo de propiedades: {e}")
            return {'error': str(e), 'cargadas': 0, 'errores': 1}

    def cargar_servicios(self, archivo_servicios: Path) -> Dict[str, Any]:
        """
        Cargar servicios urbanos a PostgreSQL.

        Args:
            archivo_servicios: Archivo Excel con servicios

        Returns:
            Resultado de la carga
        """
        logger.info(f"Cargando servicios desde: {archivo_servicios.name}")

        try:
            # Leer Excel
            wb = load_workbook(archivo_servicios)
            servicios_cargados = 0
            errores = []

            # Crear respaldo
            backup_file = self.crear_respaldo_tabla('servicios')

            # Procesar cada hoja de categoría
            for sheet_name in wb.sheetnames:
                if sheet_name in ['Errores', 'Estadísticas', 'Metadatos']:
                    continue

                logger.info(f"Procesando categoría de servicios: {sheet_name}")

                df = pd.read_excel(archivo_servicios, sheet_name=sheet_name, engine='openpyxl')

                for idx, row in df.iterrows():
                    try:
                        # Preparar coordenadas PostGIS
                        coordenadas_postgis = None
                        if (not pd.isna(row.get('latitud')) and not pd.isna(row.get('longitud')) and
                            row.get('coordenadas_validas', False)):
                            coordenadas_postgis = f"ST_GeographyFromText('SRID=4326;POINT({row['longitud']} {row['latitud']})')"

                        # Insertar servicio
                        if coordenadas_postgis:
                            query = f"""
                            INSERT INTO servicios (
                                nombre, tipo_servicio, subtipo_servicio, direccion, zona,
                                coordenadas, telefono, horario, fuente_datos, fecha_registro,
                                coordenadas_validas
                            ) VALUES (
                                %s, %s, %s, %s, %s, {coordenadas_postgis}, %s, %s, %s, %s, %s
                            )
                            """
                        else:
                            query = """
                            INSERT INTO servicios (
                                nombre, tipo_servicio, subtipo_servicio, direccion, zona,
                                coordenadas, telefono, horario, fuente_datos, fecha_registro,
                                coordenadas_validas
                            ) VALUES (
                                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                            )
                            """

                        params = (
                            row.get('nombre'),
                            sheet_name,  # tipo_servicio es el nombre de la hoja
                            row.get('subtipo_servicio', sheet_name),
                            row.get('direccion'),
                            row.get('zona'),
                            None,  # coordenadas se maneja aparte
                            row.get('telefono'),
                            row.get('horario'),
                            row.get('fuente_datos', 'guia_urbana_municipal'),
                            row.get('fecha_registro', datetime.now()),
                            row.get('coordenadas_validas', False)
                        )

                        with self.connection.cursor() as cursor:
                            if coordenadas_postgis:
                                cursor.execute(query, params)
                            else:
                                cursor.execute(query + (None,))
                            servicios_cargados += 1

                    except Exception as e:
                        errores.append({
                            'hoja': sheet_name,
                            'fila': idx + 1,
                            'error': str(e),
                            'datos': {
                                'nombre': row.get('nombre'),
                                'direccion': row.get('direccion')
                            }
                        })
                        logger.error(f"Error cargando servicio fila {idx+1}: {e}")

            self.stats['servicios_cargados'] = servicios_cargados

            return {
                'cargados': servicios_cargados,
                'errores': len(errores),
                'detalle_errores': errores,
                'backup_file': backup_file
            }

        except Exception as e:
            logger.error(f"Error procesando archivo de servicios: {e}")
            return {'error': str(e), 'cargados': 0, 'errores': 1}

    def validar_datos_cargados(self) -> Dict[str, Any]:
        """
        Validar datos cargados en PostgreSQL.

        Returns:
            Resultado de validación
        """
        logger.info("Validando datos cargados en PostgreSQL...")

        try:
            # Contar registros en cada tabla
            query_agentes = "SELECT COUNT(*) FROM agentes"
            query_propiedades = "SELECT COUNT(*) FROM propiedades"
            query_servicios = "SELECT COUNT(*) FROM servicios"

            with self.connection.cursor() as cursor:
                cursor.execute(query_agentes)
                count_agentes = cursor.fetchone()[0]

                cursor.execute(query_propiedades)
                count_propiedades = cursor.fetchone()[0]

                cursor.execute(query_servicios)
                count_servicios = cursor.fetchone()[0]

            # Validar coordenadas
            query_coordenadas_propiedades = """
            SELECT COUNT(*) FROM propiedades WHERE coordenadas_validas = true
            """
            query_coordenadas_servicios = """
            SELECT COUNT(*) FROM servicios WHERE coordenadas_validas = true
            """

            with self.connection.cursor() as cursor:
                cursor.execute(query_coordenadas_propiedades)
                coords_propiedades = cursor.fetchone()[0]

                cursor.execute(query_coordenadas_servicios)
                coords_servicios = cursor.fetchone()[0]

            # Verificar índices espaciales
            query_indices = """
            SELECT indexname, tablename FROM pg_indexes
            WHERE tablename IN ('agentes', 'propiedades', 'servicios')
            AND indexname LIKE '%coordenadas%'
            """

            with self.connection.cursor() as cursor:
                cursor.execute(query_indices)
                indices = cursor.fetchall()

            return {
                'conteos': {
                    'agentes': count_agentes,
                    'propiedades': count_propiedades,
                    'servicios': count_servicios
                },
                'coordenadas_validas': {
                    'propiedades': coords_propiedades,
                    'servicios': coords_servicios
                },
                'indices_espaciales': indices,
                'porcentaje_coordenadas': {
                    'propiedades': round(coords_propiedades / count_propiedades * 100, 2) if count_propiedades > 0 else 0,
                    'servicios': round(coords_servicios / count_servicios * 100, 2) if count_servicios > 0 else 0
                }
            }

        except Exception as e:
            logger.error(f"Error validando datos: {e}")
            return {'error': str(e)}

    def procesar_archivos_validados(self):
        """
        Procesar todos los archivos validados y cargarlos a PostgreSQL.
        """
        if not self.conectar_db():
            logger.error("No se pudo conectar a la base de datos")
            return

        try:
            # Buscar archivos validados
            archivos_agentes = list(self.validated_dir.glob("agentes_consolidados.xlsx"))
            archivos_propiedades = list(self.validated_dir.glob("propiedades_*_validado.xlsx"))
            archivos_servicios = list(self.validated_dir.glob("servicios_*_validado.xlsx"))

            resultados = {}

            # 1. Cargar agentes
            if archivos_agentes:
                logger.info(f"Encontrado archivo de agentes: {archivos_agentes[0].name}")
                resultados['agentes'] = self.cargar_agentes(archivos_agentes[0])
                self.stats['archivos_procesados'] += 1
            else:
                logger.warning("No se encontró archivo de agentes consolidados")
                resultados['agentes'] = {'error': 'Archivo no encontrado', 'cargados': 0}

            # 2. Cargar propiedades
            for archivo_prop in archivos_propiedades:
                logger.info(f"Procesando archivo de propiedades: {archivo_prop.name}")
                resultado = self.cargar_propiedades(archivo_prop)
                if 'propiedades' not in resultados:
                    resultados['propiedades'] = {'cargadas': 0, 'errores': 0, 'detalle_errores': []}

                resultados['propiedades']['cargadas'] += resultado.get('cargadas', 0)
                resultados['propiedades']['errores'] += resultado.get('errores', 0)
                if 'detalle_errores' in resultado:
                    resultados['propiedades']['detalle_errores'].extend(resultado['detalle_errores'])

                self.stats['archivos_procesados'] += 1

            # 3. Cargar servicios
            for archivo_serv in archivos_servicios:
                logger.info(f"Procesando archivo de servicios: {archivo_serv.name}")
                resultado = self.cargar_servicios(archivo_serv)
                if 'servicios' not in resultados:
                    resultados['servicios'] = {'cargados': 0, 'errores': 0, 'detalle_errores': []}

                resultados['servicios']['cargados'] += resultado.get('cargados', 0)
                resultados['servicios']['errores'] += resultado.get('errores', 0)
                if 'detalle_errores' in resultado:
                    resultados['servicios']['detalle_errores'].extend(resultado['detalle_errores'])

                self.stats['archivos_procesados'] += 1

            # 4. Validar datos cargados
            validacion = self.validar_datos_cargados()
            resultados['validacion'] = validacion

            # 5. Generar reporte final
            self.generar_reporte_final(resultados)

            logger.info("Proceso de carga a PostgreSQL completado")

        except Exception as e:
            logger.error(f"Error en proceso de carga: {e}")
        finally:
            self.desconectar_db()

    def generar_reporte_final(self, resultados: Dict[str, Any]):
        """
        Generar reporte final del proceso de carga.

        Args:
            resultados: Resultados del proceso de carga
        """
        reporte = {
            'resumen_carga': {
                'fecha': datetime.now().isoformat(),
                'archivos_procesados': self.stats['archivos_procesados'],
                'agentes_cargados': self.stats['agentes_cargados'],
                'propiedades_cargadas': self.stats['propiedades_cargadas'],
                'servicios_cargados': self.stats['servicios_cargados'],
                'errores_totales': self.stats['errores_carga']
            },
            'resultados_detalle': resultados,
            'validacion_datos': resultados.get('validacion', {}),
            'configuracion_db': {
                'host': self.db_config['host'],
                'database': self.db_config['database'],
                'user': self.db_config['user']
            }
        }

        # Guardar reporte
        reporte_path = self.logs_dir / f"reporte_carga_postgres_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(reporte_path, 'w', encoding='utf-8') as f:
            json.dump(reporte, f, indent=2, ensure_ascii=False)

        # Imprimir resumen
        logger.info("=" * 80)
        logger.info("REPORTE FINAL - CARGA A POSTGRESQL")
        logger.info("=" * 80)
        logger.info(f"Archivos procesados: {self.stats['archivos_procesados']}")
        logger.info(f"Agentes cargados: {self.stats['agentes_cargados']}")
        logger.info(f"Propiedades cargadas: {self.stats['propiedades_cargadas']}")
        logger.info(f"Servicios cargados: {self.stats['servicios_cargados']}")

        if 'validacion' in resultados and 'conteos' in resultados['validacion']:
            validacion = resultados['validacion']
            logger.info("")
            logger.info("Validación en PostgreSQL:")
            logger.info(f"  - Agentes en DB: {validacion['conteos']['agentes']}")
            logger.info(f"  - Propiedades en DB: {validacion['conteos']['propiedades']}")
            logger.info(f"  - Servicios en DB: {validacion['conteos']['servicios']}")

            if 'porcentaje_coordenadas' in validacion:
                logger.info("")
                logger.info("Calidad de coordenadas:")
                logger.info(f"  - Propiedades con coordenadas: {validacion['coordenadas_validas']['propiedades']} ({validacion['porcentaje_coordenadas']['propiedades']}%)")
                logger.info(f"  - Servicios con coordenadas: {validacion['coordenadas_validas']['servicios']} ({validacion['porcentaje_coordenadas']['servicios']}%)")

        logger.info("")
        logger.info(f"Reporte guardado en: {reporte_path}")
        logger.info("=" * 80)


def main():
    """Función principal para ejecutar el ETL."""
    # Configuración de directorios
    validated_dir = "data/intermedio/validados"
    logs_dir = "data/postgres/logs"
    backups_dir = "data/postgres/backups"

    # Verificar configuración de base de datos
    if not os.getenv('DB_HOST'):
        logger.error("Variables de entorno de base de datos no configuradas. Por favor configure DB_HOST, DB_NAME, DB_USER, DB_PASSWORD")
        return

    # Crear instancia y ejecutar
    etl = IntermediateToPostgresETL(validated_dir, logs_dir, backups_dir)
    etl.procesar_archivos_validados()


if __name__ == "__main__":
    main()