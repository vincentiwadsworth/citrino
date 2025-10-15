#!/usr/bin/env python3
"""
ETL Excel to Intermediate - Sprint 1 PostgreSQL Migration
=========================================================

Este script procesa archivos Excel crudos de relevamiento de propiedades
y genera archivos intermedios limpios en formato Excel para revisión humana.

Flujo: Excel Crudo → Excel Limpio → (Revisión Humana) → PostgreSQL

Author: Claude Code for Citrino
Date: October 2025
"""

import os
import re
import json
import logging
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/postgres/logs/etl_excel_to_intermediate.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ExcelToIntermediateETL:
    """ETL para procesar archivos Excel de propiedades hacia formato intermedio."""

    # Bounds de Santa Cruz de la Sierra para validación de coordenadas
    SANTA_CRUZ_BOUNDS = {
        'lat_min': -18.5,
        'lat_max': -17.5,
        'lon_min': -63.5,
        'lon_max': -63.0
    }

    # Mapeo de tipos de propiedad estándar
    TIPO_PROPIEDAD_MAP = {
        'casa': 'Casa',
        'departamento': 'Departamento',
        'depto': 'Departamento',
        'apartamento': 'Departamento',
        'terreno': 'Terreno/Lote',
        'lote': 'Terreno/Lote',
        'oficina': 'Oficina/Comercial',
        'local': 'Oficina/Comercial',
        'comercial': 'Oficina/Comercial',
        'galpon': 'Galpón/Industrial',
        'industrial': 'Galpón/Industrial'
    }

    # Mapeo de estados de propiedad
    ESTADO_PROPIEDAD_MAP = {
        'venta': 'En Venta',
        'alquiler': 'En Alquiler',
        'alquiler temporal': 'Alquiler Temporal',
        'anticrético': 'Anticrético',
        'preventa': 'Preventa',
        'proyecto': 'En Proyecto'
    }

    def __init__(self, input_dir: str, output_dir: str, errors_dir: str):
        """
        Inicializar el ETL.

        Args:
            input_dir: Directorio con archivos Excel crudos
            output_dir: Directorio para archivos procesados
            errors_dir: Directorio para logs de errores
        """
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.errors_dir = Path(errors_dir)

        # Asegurar que los directorios existan
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.errors_dir.mkdir(parents=True, exist_ok=True)

        # Estadísticas del proceso
        self.stats = {
            'total_archivos': 0,
            'total_propiedades': 0,
            'propiedades_validas': 0,
            'agentes_extraidos': 0,
            'errores_detectados': 0,
            'coordenadas_validas': 0,
            'coordenadas_invalidas': 0
        }

    def limpiar_texto(self, texto: str) -> str:
        """
        Limpiar y normalizar texto.

        Args:
            texto: Texto a limpiar

        Returns:
            Texto limpio y normalizado
        """
        if pd.isna(texto) or texto is None:
            return ""

        # Convertir a string y normalizar Unicode
        texto = str(texto)
        texto = unicodedata.normalize('NFKD', texto)

        # Quitar caracteres extraños y espacios múltiples
        texto = re.sub(r'[^\w\s.,-:/\(\)]', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto)

        return texto.strip().title()

    def extraer_coordenadas(self, texto: str) -> Optional[Tuple[float, float]]:
        """
        Extraer coordenadas latitud/longitud de un texto.

        Args:
            texto: Texto que puede contener coordenadas

        Returns:
            Tupla (latitud, longitud) o None si no encuentra
        """
        if pd.isna(texto) or texto is None:
            return None

        texto = str(texto).lower()

        # Patrones de coordenadas comunes
        patrones = [
            r'(-?\d+\.?\d*)[,\s]+(-?\d+\.?\d*)',  # lat,lon o lat lon
            r'lat[:\s]+(-?\d+\.?\d*)[,\s]+lon[:\s]+(-?\d+\.?\d*)',  # lat: X, lon: Y
            r'[-+]?\d{1,3}\.\d{6}[,/\s]+[-+]?\d{1,3}\.\d{6}',  # Coordenadas decimales
        ]

        for patron in patrones:
            match = re.search(patron, texto)
            if match:
                try:
                    lat = float(match.group(1))
                    lon = float(match.group(2))

                    # Validar bounds de Santa Cruz
                    if (self.SANTA_CRUZ_BOUNDS['lat_min'] <= lat <= self.SANTA_CRUZ_BOUNDS['lat_max'] and
                        self.SANTA_CRUZ_BOUNDS['lon_min'] <= lon <= self.SANTA_CRUZ_BOUNDS['lon_max']):
                        return (lat, lon)
                except ValueError:
                    continue

        return None

    def normalizar_tipo_propiedad(self, texto: str) -> str:
        """
        Normalizar tipo de propiedad.

        Args:
            texto: Texto del tipo de propiedad

        Returns:
            Tipo de propiedad normalizado
        """
        if pd.isna(texto) or texto is None:
            return "No especificado"

        texto = str(texto).lower().strip()

        for clave, valor in self.TIPO_PROPIEDAD_MAP.items():
            if clave in texto:
                return valor

        return texto.title()

    def extraer_agente(self, texto: str) -> Optional[Dict[str, str]]:
        """
        Extraer información del agente inmobiliario.

        Args:
            texto: Texto con información del agente

        Returns:
            Diccionario con datos del agente o None
        """
        if pd.isna(texto) or texto is None:
            return None

        texto = str(texto)

        # Patrones para extraer información de contacto
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        phone_pattern = r'(\+?591[-\s]?|0)?(\d{8}|\d{4}[-\s]?\d{4})'

        email_match = re.search(email_pattern, texto)
        phone_match = re.search(phone_pattern, texto)

        # Extraer nombre (asumir primera parte del texto)
        nombre = self.limpiar_texto(texto.split(',')[0])

        if nombre and len(nombre) > 2:
            return {
                'nombre': nombre,
                'telefono': phone_match.group(0) if phone_match else "",
                'email': email_match.group(0) if email_match else "",
                'empresa': self.limpiar_texto(texto.split(',')[-1]) if ',' in texto else ""
            }

        return None

    def validar_precio(self, texto: str) -> Optional[float]:
        """
        Validar y extraer precio en USD.

        Args:
            texto: Texto con precio

        Returns:
            Precio en USD o None si no es válido
        """
        if pd.isna(texto) or texto is None:
            return None

        texto = str(texto).lower().replace('$', '').replace(',', '').replace('.', '')

        # Buscar números
        numeros = re.findall(r'\d+', texto)
        if numeros:
            try:
                precio = float(numeros[0])

                # Validar rangos razonables para Santa Cruz
                if 1000 <= precio <= 10000000:
                    return precio
            except ValueError:
                pass

        return None

    def procesar_archivo_excel(self, archivo_path: Path) -> Dict[str, Any]:
        """
        Procesar un archivo Excel de propiedades.

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Diccionario con resultados del procesamiento
        """
        logger.info(f"Procesando archivo: {archivo_path.name}")

        try:
            # Leer Excel
            df = pd.read_excel(archivo_path, engine='openpyxl')

            # Estandarizar nombres de columnas
            df.columns = [col.lower().replace(' ', '_') for col in df.columns]

            # Listas para datos procesados
            propiedades_procesadas = []
            agentes_extraidos = []
            errores_detectados = []

            for idx, row in df.iterrows():
                try:
                    # Extraer coordenadas
                    coordenadas = None
                    for col in ['coordenadas', 'ubicacion', 'direccion', 'descripcion']:
                        if col in df.columns and not pd.isna(row[col]):
                            coordenadas = self.extraer_coordenadas(str(row[col]))
                            if coordenadas:
                                break

                    # Extraer información del agente
                    agente_info = None
                    for col in ['contacto', 'agente', 'telefono', 'email']:
                        if col in df.columns and not pd.isna(row[col]):
                            agente_info = self.extraer_agente(str(row[col]))
                            if agente_info:
                                agente_info['fuente_archivo'] = archivo_path.name
                                agentes_extraidos.append(agente_info)
                                break

                    # Validar precio
                    precio = None
                    for col in ['precio', 'precio_usd', 'valor', 'costo']:
                        if col in df.columns and not pd.isna(row[col]):
                            precio = self.validar_precio(str(row[col]))
                            break

                    # Construir propiedad procesada
                    propiedad = {
                        'id_original': idx + 1,
                        'titulo': self.limpiar_texto(row.get('titulo', row.get('descripcion', ''))),
                        'descripcion': self.limpiar_texto(row.get('descripcion', '')),
                        'tipo_propiedad': self.normalizar_tipo_propiedad(
                            row.get('tipo', row.get('categoria', ''))
                        ),
                        'estado_propiedad': 'En Venta',  # Default
                        'precio_usd': precio,
                        'direccion': self.limpiar_texto(row.get('direccion', '')),
                        'zona': self.limpiar_texto(row.get('zona', row.get('sector', ''))),
                        'superficie_total': self._extraer_numero(row.get('superficie', row.get('area', ''))),
                        'superficie_construida': self._extraer_numero(row.get('superficie_construida', '')),
                        'num_dormitorios': self._extraer_numero(row.get('dormitorios', row.get('habitaciones', '')), entero=True),
                        'num_banos': self._extraer_numero(row.get('banos', row.get('baños', '')), entero=True),
                        'num_garajes': self._extraer_numero(row.get('garajes', row.get('estacionamiento', '')), entero=True),
                        'latitud': coordenadas[0] if coordenadas else None,
                        'longitud': coordenadas[1] if coordenadas else None,
                        'coordenadas_validas': coordenadas is not None,
                        'url_origen': row.get('url', ''),
                        'fecha_scraping': datetime.now(),
                        'proveedor_datos': 'excel_crudo',
                        'codigo_proveedor': f"{archivo_path.stem}_{idx+1}"
                    }

                    # Validaciones
                    if not propiedad['titulo'] or len(propiedad['titulo']) < 5:
                        errores_detectados.append({
                            'fila': idx + 1,
                            'tipo_error': 'Título inválido',
                            'descripcion': 'El título es demasiado corto o está vacío',
                            'sugerencia': 'Revisar título de la propiedad',
                            'severidad': 'Alta'
                        })

                    if not propiedad['precio_usd']:
                        errores_detectados.append({
                            'fila': idx + 1,
                            'tipo_error': 'Precio inválido',
                            'descripcion': 'No se pudo extraer un precio válido',
                            'sugerencia': 'Verificar formato del precio',
                            'severidad': 'Media'
                        })

                    propiedades_procesadas.append(propiedad)

                except Exception as e:
                    errores_detectados.append({
                        'fila': idx + 1,
                        'tipo_error': 'Error procesamiento',
                        'descripcion': str(e),
                        'sugerencia': 'Revisar datos de la fila',
                        'severidad': 'Alta'
                    })
                    logger.error(f"Error procesando fila {idx+1}: {e}")

            # Actualizar estadísticas
            self.stats['total_propiedades'] += len(propiedades_procesadas)
            self.stats['agentes_extraidos'] += len(agentes_extraidos)
            self.stats['errores_detectados'] += len(errores_detectados)
            self.stats['coordenadas_validas'] += sum(1 for p in propiedades_procesadas if p['coordenadas_validas'])
            self.stats['coordenadas_invalidas'] += sum(1 for p in propiedades_procesadas if not p['coordenadas_validas'])
            self.stats['propiedades_validas'] += len([p for p in propiedades_procesadas if p['titulo'] and p['precio_usd']])

            return {
                'propiedades': propiedades_procesadas,
                'agentes': agentes_extraidos,
                'errores': errores_detectados,
                'stats': {
                    'total_filas': len(df),
                    'propiedades_procesadas': len(propiedades_procesadas),
                    'agentes_extraidos': len(agentes_extraidos),
                    'errores_encontrados': len(errores_detectados),
                    'coordenadas_validas': sum(1 for p in propiedades_procesadas if p['coordenadas_validas']),
                    'coordenadas_invalidas': sum(1 for p in propiedades_procesadas if not p['coordenadas_validas']),
                    'porcentaje_calidad': round(len([p for p in propiedades_procesadas if p['titulo'] and p['precio_usd']]) / len(propiedades_procesadas) * 100, 2) if propiedades_procesadas else 0
                }
            }

        except Exception as e:
            logger.error(f"Error procesando archivo {archivo_path}: {e}")
            return {
                'propiedades': [],
                'agentes': [],
                'errores': [{'fila': 0, 'tipo_error': 'Error archivo', 'descripcion': str(e), 'sugerencia': 'Revisar formato del archivo', 'severidad': 'Crítico'}],
                'stats': {'error': str(e)}
            }

    def _extraer_numero(self, texto: str, entero: bool = False) -> Optional[float]:
        """
        Extraer número de un texto.

        Args:
            texto: Texto que contiene un número
            entero: Si es True, devuelve entero

        Returns:
            Número extraído o None
        """
        if pd.isna(texto) or texto is None:
            return None

        numeros = re.findall(r'\d+\.?\d*', str(texto))
        if numeros:
            try:
                numero = float(numeros[0])
                return int(numero) if entero else numero
            except ValueError:
                pass

        return None

    def guardar_archivo_intermedio(self, resultados: Dict[str, Any], archivo_original: str):
        """
        Guardar resultados en archivo Excel intermedio.

        Args:
            resultados: Datos procesados
            archivo_original: Nombre del archivo original
        """
        output_filename = f"propiedades_{archivo_original}_procesado.xlsx"
        output_path = self.output_dir / output_filename

        # Crear workbook con múltiples hojas
        wb = Workbook()

        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # 1. Hoja de Propiedades
        if resultados['propiedades']:
            ws_propiedades = wb.create_sheet("Propiedades")

            # Convertir a DataFrame para mejor manejo
            df_propiedades = pd.DataFrame(resultados['propiedades'])

            # Escribir encabezados
            for r in dataframe_to_rows(df_propiedades, index=True, header=True):
                ws_propiedades.append(r)

            # Formato de encabezados
            for cell in ws_propiedades[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # 2. Hoja de Agentes
        if resultados['agentes']:
            ws_agentes = wb.create_sheet("Agentes")
            df_agentes = pd.DataFrame(resultados['agentes'])

            # Eliminar duplicados por nombre/email
            df_agentes = df_agentes.drop_duplicates(subset=['nombre', 'email'], keep='first')

            for r in dataframe_to_rows(df_agentes, index=False, header=True):
                ws_agentes.append(r)

            # Formato de encabezados
            for cell in ws_agentes[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

        # 3. Hoja de Errores
        if resultados['errores']:
            ws_errores = wb.create_sheet("Errores")
            df_errores = pd.DataFrame(resultados['errores'])

            for r in dataframe_to_rows(df_errores, index=False, header=True):
                ws_errores.append(r)

            # Formato de encabezados
            for cell in ws_errores[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

            # Colorear filas según severidad
            for row in ws_errores.iter_rows(min_row=2, max_row=ws_errores.max_row, min_col=5, max_col=5):
                for cell in row:
                    if cell.value == 'Crítico':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif cell.value == 'Alta':
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    elif cell.value == 'Media':
                        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")

        # 4. Hoja de Estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        stats_data = [
            ['Métrica', 'Valor'],
            ['Total de propiedades procesadas', resultados['stats'].get('propiedades_procesadas', 0)],
            ['Coordenadas válidas', resultados['stats'].get('coordenadas_validas', 0)],
            ['Coordenadas inválidas', resultados['stats'].get('coordenadas_invalidas', 0)],
            ['Agentes extraídos', resultados['stats'].get('agentes_extraidos', 0)],
            ['Errores encontrados', resultados['stats'].get('errores_encontrados', 0)],
            ['Porcentaje de calidad', f"{resultados['stats'].get('porcentaje_calidad', 0)}%"],
            ['Fecha de procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Archivo original', archivo_original],
            ['Versión del script', '1.0.0']
        ]

        for row in stats_data:
            ws_stats.append(row)

        # Formato de encabezados y estadísticas
        for cell in ws_stats[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 5. Hoja de Metadatos
        ws_metadatos = wb.create_sheet("Metadatos")
        metadatos_data = [
            ['Propiedad', 'Valor'],
            ['Archivo origen', archivo_original],
            ['Fecha procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Tiempo procesamiento', 'N/A'],  # Se llenaría después
            ['Versión script', '1.0.0'],
            ['Responsable', 'ETL Automático'],
            ['Bounds validación', f"Lat: {self.SANTA_CRUZ_BOUNDS['lat_min']} a {self.SANTA_CRUZ_BOUNDS['lat_max']}, Lon: {self.SANTA_CRUZ_BOUNDS['lon_min']} a {self.SANTA_CRUZ_BOUNDS['lon_max']}"]
        ]

        for row in metadatos_data:
            ws_metadatos.append(row)

        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Archivo intermedio guardado: {output_path}")

        return output_path

    def procesar_todos_los_archivos(self):
        """
        Procesar todos los archivos Excel en el directorio de entrada.
        """
        archivos_excel = list(self.input_dir.glob("*.xlsx")) + list(self.input_dir.glob("*.xls"))

        if not archivos_excel:
            logger.warning(f"No se encontraron archivos Excel en {self.input_dir}")
            return

        self.stats['total_archivos'] = len(archivos_excel)
        logger.info(f"Encontrados {len(archivos_excel)} archivos Excel para procesar")

        for archivo_path in archivos_excel:
            try:
                # Procesar archivo
                resultados = self.procesar_archivo_excel(archivo_path)

                # Guardar archivo intermedio
                self.guardar_archivo_intermedio(resultados, archivo_path.stem)

                # Guardar log de errores
                if resultados['errores']:
                    errores_path = self.errors_dir / f"errores_{archivo_path.stem}.json"
                    with open(errores_path, 'w', encoding='utf-8') as f:
                        json.dump(resultados['errores'], f, indent=2, ensure_ascii=False)

            except Exception as e:
                logger.error(f"Error procesando archivo {archivo_path}: {e}")
                continue

        # Generar reporte final
        self.generar_reporte_final()

    def generar_reporte_final(self):
        """
        Generar reporte final del procesamiento.
        """
        reporte = {
            'resumen_ejecucion': {
                'fecha': datetime.now().isoformat(),
                'total_archivos': self.stats['total_archivos'],
                'total_propiedades': self.stats['total_propiedades'],
                'propiedades_validas': self.stats['propiedades_validas'],
                'agentes_extraidos': self.stats['agentes_extraidos'],
                'errores_detectados': self.stats['errores_detectados'],
                'coordenadas_validas': self.stats['coordenadas_validas'],
                'coordenadas_invalidas': self.stats['coordenadas_invalidas']
            },
            'metricas_calidad': {
                'porcentaje_propiedades_validas': round(self.stats['propiedades_validas'] / self.stats['total_propiedades'] * 100, 2) if self.stats['total_propiedades'] > 0 else 0,
                'porcentaje_coordenadas_validas': round(self.stats['coordenadas_validas'] / self.stats['total_propiedades'] * 100, 2) if self.stats['total_propiedades'] > 0 else 0,
                'promedio_agentes_por_archivo': round(self.stats['agentes_extraidos'] / self.stats['total_archivos'], 2) if self.stats['total_archivos'] > 0 else 0
            }
        }

        # Guardar reporte
        reporte_path = self.errors_dir / f"reporte_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(reporte_path, 'w', encoding='utf-8') as f:
            json.dump(reporte, f, indent=2, ensure_ascii=False)

        logger.info("=" * 80)
        logger.info("REPORTE FINAL DE PROCESAMIENTO")
        logger.info("=" * 80)
        logger.info(f"Archivos procesados: {self.stats['total_archivos']}")
        logger.info(f"Total propiedades: {self.stats['total_propiedades']}")
        logger.info(f"Propiedades válidas: {self.stats['propiedades_validas']} ({reporte['metricas_calidad']['porcentaje_propiedades_validas']}%)")
        logger.info(f"Coordenadas válidas: {self.stats['coordenadas_validas']} ({reporte['metricas_calidad']['porcentaje_coordenadas_validas']}%)")
        logger.info(f"Agentes extraídos: {self.stats['agentes_extraídos']}")
        logger.info(f"Errores detectados: {self.stats['errores_detectados']}")
        logger.info(f"Reporte guardado en: {reporte_path}")
        logger.info("=" * 80)


def main():
    """Función principal para ejecutar el ETL."""
    # Configuración de directorios
    input_dir = "data/raw/relevamiento"
    output_dir = "data/intermedio/procesados"
    errors_dir = "data/intermedio/errores"

    # Crear instancia y ejecutar
    etl = ExcelToIntermediateETL(input_dir, output_dir, errors_dir)
    etl.procesar_todos_los_archivos()


if __name__ == "__main__":
    main()