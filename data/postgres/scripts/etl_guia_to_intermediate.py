#!/usr/bin/env python3
"""
ETL Guía Urbana to Intermediate - Sprint 1 PostgreSQL Migration
===============================================================

Este script procesa el archivo Excel de la guía urbana municipal
y genera archivos intermedios organizados por tipo de servicio.

Flujo: Guía Urbana Excel → Servicios Clasificados → (Revisión Humana) → PostgreSQL

Author: Claude Code for Citrino
Date: October 2025
"""

import os
import re
import json
import logging
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/postgres/logs/etl_guia_to_intermediate.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class GuiaToIntermediateETL:
    """ETL para procesar guía urbana municipal hacia formato intermedio."""

    # Bounds de Santa Cruz de la Sierra para validación de coordenadas
    SANTA_CRUZ_BOUNDS = {
        'lat_min': -18.5,
        'lat_max': -17.5,
        'lon_min': -63.5,
        'lon_max': -63.0
    }

    # Clasificación de servicios por palabras clave
    CLASIFICACION_SERVICIOS = {
        'Educación': {
            'palabras_clave': ['colegio', 'escuela', 'universidad', 'instituto', 'academia', 'centro educativo', 'colegio nacional', 'unidad educativa', 'kínder', 'jardín', 'guardería', 'biblioteca', 'educación', 'learning', 'school', 'college', 'university'],
            'subtipos': {
                'Colegio': ['colegio', 'escuela', 'unidad educativa', 'kínder', 'jardín', 'guardería'],
                'Universidad': ['universidad', 'universitario', 'facultad', 'instituto superior'],
                'Academia': ['academia', 'instituto', 'centro de estudios', 'cursos'],
                'Biblioteca': ['biblioteca', 'centro de lectura']
            }
        },
        'Salud': {
            'palabras_clave': ['hospital', 'clínica', 'farmacia', 'consultorio', 'médico', 'salud', 'centro médico', 'laboratorio', 'odontólogo', 'dentista', 'pediatra', 'cardiólogo', 'médico', 'doctor', 'clinica', 'hospital', 'pharmacy', 'medical', 'health'],
            'subtipos': {
                'Hospital': ['hospital', 'centro médico', 'centro de salud'],
                'Clínica': ['clínica', 'consultorio', 'centro médico especializado'],
                'Farmacia': ['farmacia', 'botica', 'droguería', 'pharmacy'],
                'Laboratorio': ['laboratorio', 'laboratorio clínico', 'análisis clínico']
            }
        },
        'Comercio': {
            'palabras_clave': ['supermercado', 'tienda', 'mercado', 'comercio', 'shopping', 'centro comercial', 'store', 'shop', 'market', 'supermarket', 'comercial', 'venta', 'negocio', 'bodega', 'abarrotes', 'almacén'],
            'subtipos': {
                'Supermercado': ['supermercado', 'hipermercado', 'supermarket'],
                'Tienda': ['tienda', 'store', 'shop', 'negocio', 'bodega', 'abarrotes'],
                'Mercado': ['mercado', 'mercado municipal', 'plaza de mercado'],
                'Centro Comercial': ['shopping', 'centro comercial', 'mall']
            }
        },
        'Servicios': {
            'palabras_clave': ['banco', 'oficina', 'gobierno', 'municipal', 'servicio', 'institución', 'oficina pública', 'banco', 'cajero', 'atm', 'financiera', 'cooperativa', 'entidad', 'organismo', 'ministerio', 'gobierno municipal'],
            'subtipos': {
                'Banco': ['banco', 'cajero', 'atm', 'financiera', 'cooperativa', 'entidad bancaria'],
                'Oficina Gubernamental': ['municipal', 'gobierno', 'ministerio', 'organismo público', 'institución'],
                'Oficina Privada': ['oficina', 'servicio privado', 'consultora', 'asesoría']
            }
        },
        'Transporte': {
            'palabras_clave': ['terminal', 'bus', 'transporte', 'taxi', 'parada', 'terminal de buses', 'estación', 'metrobus', 'transporte público', 'bus', 'taxi', 'terminal', 'transport'],
            'subtipos': {
                'Terminal': ['terminal', 'terminal de buses', 'estación terminal'],
                'Parada': ['parada', 'parada de bus', 'paradero'],
                'Servicio Taxi': ['taxi', 'radio taxi', 'movilidad']
            }
        },
        'Recreación': {
            'palabras_clave': ['parque', 'plaza', 'deporte', 'gimnasio', 'centro deportivo', 'club', 'recreación', 'área verde', 'plazoleta', 'gimnasio', 'gym', 'deportivo', 'centro recreativo', 'park', 'square', 'sports', 'gym'],
            'subtipos': {
                'Parque': ['parque', 'área verde', 'plaza', 'plazoleta', 'park'],
                'Centro Deportivo': ['centro deportivo', 'gimnasio', 'gym', 'club deportivo'],
                'Recreación': ['centro recreativo', 'área de juegos', 'recreación']
            }
        }
    }

    def __init__(self, input_dir: str, output_dir: str, errors_dir: str):
        """
        Inicializar el ETL.

        Args:
            input_dir: Directorio con archivo de guía urbana
            output_dir: Directorio para archivos procesados
            errors_dir: Directorio para logs de errores
        """
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.errors_dir = Path(errors_dir)

        # Asegurar que los directorios existan
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.errors_dir.mkdir(parents=True, exist_ok=True)

        # Estadísticas del proceso
        self.stats = {
            'total_servicios': 0,
            'servicios_clasificados': 0,
            'servicios_con_coordenadas': 0,
            'por_categoria': {},
            'errores_detectados': 0
        }

    def limpiar_texto(self, texto: str) -> str:
        """
        Limpiar y normalizar texto.

        Args:
            texto: Texto a limpiar

        Returns:
            Texto limpio y normalizado
        """
        if pd.isna(texto) or texto is None:
            return ""

        # Convertir a string y normalizar Unicode
        texto = str(texto)
        texto = unicodedata.normalize('NFKD', texto)

        # Quitar caracteres extraños y espacios múltiples
        texto = re.sub(r'[^\w\s.,-:/()\'°]', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto)

        return texto.strip().title()

    def extraer_coordenadas(self, texto: str) -> Optional[Tuple[float, float]]:
        """
        Extraer coordenadas latitud/longitud de un texto.

        Args:
            texto: Texto que puede contener coordenadas

        Returns:
            Tupla (latitud, longitud) o None si no encuentra
        """
        if pd.isna(texto) or texto is None:
            return None

        texto = str(texto).lower()

        # Patrones de coordenadas comunes
        patrones = [
            r'(-?\d+\.?\d*)[,\s]+(-?\d+\.?\d*)',  # lat,lon o lat lon
            r'lat[:\s]+(-?\d+\.?\d*)[,\s]+lon[:\s]+(-?\d+\.?\d*)',  # lat: X, lon: Y
            r'[-+]?\d{1,3}\.\d{6}[,/\s]+[-+]?\d{1,3}\.\d{6}',  # Coordenadas decimales
        ]

        for patron in patrones:
            match = re.search(patron, texto)
            if match:
                try:
                    lat = float(match.group(1))
                    lon = float(match.group(2))

                    # Validar bounds de Santa Cruz
                    if (self.SANTA_CRUZ_BOUNDS['lat_min'] <= lat <= self.SANTA_CRUZ_BOUNDS['lat_max'] and
                        self.SANTA_CRUZ_BOUNDS['lon_min'] <= lon <= self.SANTA_CRUZ_BOUNDS['lon_max']):
                        return (lat, lon)
                except ValueError:
                    continue

        return None

    def clasificar_servicio(self, nombre: str, descripcion: str = "") -> Tuple[str, str]:
        """
        Clasificar un servicio según su nombre y descripción.

        Args:
            nombre: Nombre del servicio
            descripcion: Descripción adicional

        Returns:
            Tupla (tipo_servicio, subtipo_servicio)
        """
        texto_completo = f"{nombre} {descripcion}".lower()

        for tipo_servicio, config in self.CLASIFICACION_SERVICIOS.items():
            # Verificar palabras clave del tipo
            for palabra in config['palabras_clave']:
                if palabra in texto_completo:
                    # Intentar encontrar subtipo
                    for subtipo, palabras_subtipo in config['subtipos'].items():
                        for palabra_sub in palabras_subtipo:
                            if palabra_sub in texto_completo:
                                return (tipo_servicio, subtipo)
                    return (tipo_servicio, tipo_servicio)  # Subtipo por defecto

        return ('Otros', 'No Clasificado')

    def extraer_telefono(self, texto: str) -> Optional[str]:
        """
        Extraer número de teléfono de un texto.

        Args:
            texto: Texto que puede contener teléfono

        Returns:
            Teléfono formateado o None
        """
        if pd.isna(texto) or texto is None:
            return None

        texto = str(texto)

        # Patrones de teléfono boliviano
        patrones = [
            r'(\+?591[-\s]?)?(\d{8})',  # +591 12345678 o 12345678
            r'(\+?591[-\s]?)?(\d{4}[-\s]?\d{4})',  # +591 1234-5678
            r'(\d{3}[-\s]?\d{3}[-\s]?\d{3})',  # 123-456-789
        ]

        for patron in patrones:
            match = re.search(patron, texto)
            if match:
                # Extraer solo los dígitos
                digitos = re.findall(r'\d', match.group(0))
                if len(digitos) == 8:
                    return f"{'-'.join([''.join(digitos[:4]), ''.join(digitos[4:])])}"
                elif len(digitos) == 7:
                    return ''.join(digitos)

        return None

    def procesar_archivo_guia(self, archivo_path: Path) -> Dict[str, Any]:
        """
        Procesar el archivo Excel de la guía urbana.

        Args:
            archivo_path: Ruta al archivo Excel

        Returns:
            Diccionario con resultados del procesamiento
        """
        logger.info(f"Procesando guía urbana: {archivo_path.name}")

        try:
            # Leer Excel
            df = pd.read_excel(archivo_path, engine='openpyxl')

            # Estandarizar nombres de columnas
            df.columns = [col.lower().replace(' ', '_') for col in df.columns]

            # Organizar servicios por categoría
            servicios_por_categoria = {categoria: [] for categoria in self.CLASIFICACION_SERVICIOS.keys()}
            servicios_por_categoria['Otros'] = []  # Para no clasificados
            errores_detectados = []

            for idx, row in df.iterrows():
                try:
                    # Extraer información básica
                    nombre = self.limpiar_texto(row.get('nombre', row.get('establecimiento', row.get('servicio', ''))))
                    direccion = self.limpiar_texto(row.get('direccion', row.get('ubicacion', '')))
                    zona = self.limpiar_texto(row.get('zona', row.get('sector', row.get('barrio', ''))))
                    descripcion = self.limpiar_texto(row.get('descripcion', row.get('actividad', '')))

                    # Extraer coordenadas
                    coordenadas = None
                    for col in ['coordenadas', 'ubicacion', 'direccion', 'descripcion']:
                        if col in df.columns and not pd.isna(row[col]):
                            coordenadas = self.extraer_coordenadas(str(row[col]))
                            if coordenadas:
                                break

                    # Clasificar servicio
                    tipo_servicio, subtipo_servicio = self.clasificar_servicio(nombre, descripcion)

                    # Extraer teléfono
                    telefono = None
                    for col in ['telefono', 'teléfono', 'contacto', 'celular']:
                        if col in df.columns and not pd.isna(row[col]):
                            telefono = self.extraer_telefono(str(row[col]))
                            if telefono:
                                break

                    # Extraer horario
                    horario = ""
                    for col in ['horario', 'horario_atencion', 'horarios']:
                        if col in df.columns and not pd.isna(row[col]):
                            horario = self.limpiar_texto(str(row[col]))
                            break

                    # Construir servicio
                    servicio = {
                        'nombre': nombre,
                        'tipo_servicio': tipo_servicio,
                        'subtipo_servicio': subtipo_servicio,
                        'direccion': direccion,
                        'zona': zona,
                        'latitud': coordenadas[0] if coordenadas else None,
                        'longitud': coordenadas[1] if coordenadas else None,
                        'coordenadas_validas': coordenadas is not None,
                        'telefono': telefono,
                        'horario': horario,
                        'fuente_datos': 'guia_urbana_municipal',
                        'fecha_registro': datetime.now()
                    }

                    # Validaciones
                    if not nombre or len(nombre) < 3:
                        errores_detectados.append({
                            'fila': idx + 1,
                            'tipo_error': 'Nombre inválido',
                            'descripcion': 'El nombre del servicio es demasiado corto o está vacío',
                            'sugerencia': 'Revisar nombre del establecimiento',
                            'severidad': 'Alta'
                        })

                    # Agregar a categoría correspondiente
                    if tipo_servicio in servicios_por_categoria:
                        servicios_por_categoria[tipo_servicio].append(servicio)
                    else:
                        servicios_por_categoria['Otros'].append(servicio)

                    # Actualizar estadísticas
                    self.stats['total_servicios'] += 1
                    if coordenadas:
                        self.stats['servicios_con_coordenadas'] += 1
                    if tipo_servicio != 'Otros':
                        self.stats['servicios_clasificados'] += 1

                    # Actualizar contador por categoría
                    if tipo_servicio not in self.stats['por_categoria']:
                        self.stats['por_categoria'][tipo_servicio] = 0
                    self.stats['por_categoria'][tipo_servicio] += 1

                except Exception as e:
                    errores_detectados.append({
                        'fila': idx + 1,
                        'tipo_error': 'Error procesamiento',
                        'descripcion': str(e),
                        'sugerencia': 'Revisar datos de la fila',
                        'severidad': 'Alta'
                    })
                    logger.error(f"Error procesando fila {idx+1}: {e}")

            self.stats['errores_detectados'] += len(errores_detectados)

            return {
                'servicios_por_categoria': servicios_por_categoria,
                'errores': errores_detectados,
                'stats': {
                    'total_servicios': self.stats['total_servicios'],
                    'servicios_clasificados': self.stats['servicios_clasificados'],
                    'servicios_con_coordenadas': self.stats['servicios_con_coordenadas'],
                    'errores_encontrados': len(errores_detectados),
                    'por_categoria': self.stats['por_categoria']
                }
            }

        except Exception as e:
            logger.error(f"Error procesando archivo {archivo_path}: {e}")
            return {
                'servicios_por_categoria': {},
                'errores': [{'fila': 0, 'tipo_error': 'Error archivo', 'descripcion': str(e), 'sugerencia': 'Revisar formato del archivo', 'severidad': 'Crítico'}],
                'stats': {'error': str(e)}
            }

    def guardar_archivo_intermedio(self, resultados: Dict[str, Any], archivo_original: str):
        """
        Guardar resultados en archivo Excel intermedio.

        Args:
            resultados: Datos procesados por categoría
            archivo_original: Nombre del archivo original
        """
        output_filename = f"servicios_urbanos_{archivo_original}_procesado.xlsx"
        output_path = self.output_dir / output_filename

        # Crear workbook con múltiples hojas por categoría
        wb = Workbook()

        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # Colores para diferentes categorías
        colores_categorias = {
            'Educación': 'E6F3FF',
            'Salud': 'FFE6E6',
            'Comercio': 'E6FFE6',
            'Servicios': 'FFF0E6',
            'Transporte': 'F0E6FF',
            'Recreación': 'E6FFFF',
            'Otros': 'F5F5F5'
        }

        # 1. Hojas por categoría de servicio
        for categoria, servicios in resultados['servicios_por_categoria'].items():
            if servicios:
                ws = wb.create_sheet(categoria)
                df_servicios = pd.DataFrame(servicios)

                # Reordenar columnas para mejor visualización
                columnas_orden = ['nombre', 'subtipo_servicio', 'direccion', 'zona', 'telefono', 'horario', 'coordenadas_validas', 'latitud', 'longitud', 'fuente_datos']
                df_servicios = df_servicios.reindex(columns=columnas_orden)

                for r in dataframe_to_rows(df_servicios, index=False, header=True):
                    ws.append(r)

                # Formato de encabezados
                color_categoria = colores_categorias.get(categoria, 'F5F5F5')
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color=color_categoria, end_color=color_categoria, fill_type="solid")

                # Autoajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Hoja de Errores
        if resultados['errores']:
            ws_errores = wb.create_sheet("Errores")
            df_errores = pd.DataFrame(resultados['errores'])

            for r in dataframe_to_rows(df_errores, index=False, header=True):
                ws_errores.append(r)

            # Formato de encabezados
            for cell in ws_errores[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # 3. Hoja de Estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        stats_data = [
            ['Métrica', 'Valor'],
            ['Total de servicios', resultados['stats'].get('total_servicios', 0)],
            ['Servicios clasificados', resultados['stats'].get('servicios_clasificados', 0)],
            ['Servicios con coordenadas', resultados['stats'].get('servicios_con_coordenadas', 0)],
            ['Errores encontrados', resultados['stats'].get('errores_encontrados', 0)],
            ['Fecha de procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Archivo original', archivo_original],
            ['Versión del script', '1.0.0']
        ]

        # Agregar estadísticas por categoría
        for categoria, cantidad in resultados['stats'].get('por_categoria', {}).items():
            stats_data.append([f'Servicios - {categoria}', cantidad])

        for row in stats_data:
            ws_stats.append(row)

        # Formato de encabezados
        for cell in ws_stats[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 4. Hoja de Metadatos
        ws_metadatos = wb.create_sheet("Metadatos")
        metadatos_data = [
            ['Propiedad', 'Valor'],
            ['Archivo origen', archivo_original],
            ['Fecha procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Versión script', '1.0.0'],
            ['Bounds validación', f"Lat: {self.SANTA_CRUZ_BOUNDS['lat_min']} a {self.SANTA_CRUZ_BOUNDS['lat_max']}, Lon: {self.SANTA_CRUZ_BOUNDS['lon_min']} a {self.SANTA_CRUZ_BOUNDS['lon_max']}"],
            ['Categorías procesadas', ', '.join(resultados['servicios_por_categoria'].keys())]
        ]

        for row in metadatos_data:
            ws_metadatos.append(row)

        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Archivo intermedio de servicios guardado: {output_path}")

        return output_path

    def procesar_guia_urbana(self):
        """
        Procesar el archivo de guía urbana municipal.
        """
        # Buscar archivo de guía urbana
        guia_files = list(self.input_dir.glob("*GUIA*.xlsx")) + list(self.input_dir.glob("*gui*a*.xlsx"))

        if not guia_files:
            logger.warning(f"No se encontró archivo de guía urbana en {self.input_dir}")
            return

        # Procesar el primer archivo encontrado
        guia_path = guia_files[0]
        logger.info(f"Procesando guía urbana: {guia_path.name}")

        # Procesar archivo
        resultados = self.procesar_archivo_guia(guia_path)

        # Guardar archivo intermedio
        self.guardar_archivo_intermedio(resultados, guia_path.stem)

        # Guardar log de errores
        if resultados['errores']:
            errores_path = self.errors_dir / f"errores_guia_{guia_path.stem}.json"
            with open(errores_path, 'w', encoding='utf-8') as f:
                json.dump(resultados['errores'], f, indent=2, ensure_ascii=False)

        # Generar reporte final
        self.generar_reporte_final(resultados)

    def generar_reporte_final(self, resultados: Dict[str, Any]):
        """
        Generar reporte final del procesamiento.
        """
        reporte = {
            'resumen_procesamiento': {
                'fecha': datetime.now().isoformat(),
                'total_servicios': self.stats['total_servicios'],
                'servicios_clasificados': self.stats['servicios_clasificados'],
                'servicios_con_coordenadas': self.stats['servicios_con_coordenadas'],
                'errores_detectados': self.stats['errores_detectados']
            },
            'distribucion_por_categoria': self.stats['por_categoria'],
            'metricas_calidad': {
                'porcentaje_clasificacion': round(self.stats['servicios_clasificados'] / self.stats['total_servicios'] * 100, 2) if self.stats['total_servicios'] > 0 else 0,
                'porcentaje_coordenadas': round(self.stats['servicios_con_coordenadas'] / self.stats['total_servicios'] * 100, 2) if self.stats['total_servicios'] > 0 else 0
            }
        }

        # Guardar reporte
        reporte_path = self.errors_dir / f"reporte_guia_urbana_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(reporte_path, 'w', encoding='utf-8') as f:
            json.dump(reporte, f, indent=2, ensure_ascii=False)

        logger.info("=" * 80)
        logger.info("REPORTE FINAL - GUÍA URBANA MUNICIPAL")
        logger.info("=" * 80)
        logger.info(f"Total servicios procesados: {self.stats['total_servicios']}")
        logger.info(f"Servicios clasificados: {self.stats['servicios_clasificados']} ({reporte['metricas_calidad']['porcentaje_clasificacion']}%)")
        logger.info(f"Servicios con coordenadas: {self.stats['servicios_con_coordenadas']} ({reporte['metricas_calidad']['porcentaje_coordenadas']}%)")
        logger.info("Distribución por categoría:")
        for categoria, cantidad in self.stats['por_categoria'].items():
            logger.info(f"  - {categoria}: {cantidad}")
        logger.info(f"Errores detectados: {self.stats['errores_detectados']}")
        logger.info(f"Reporte guardado en: {reporte_path}")
        logger.info("=" * 80)


def main():
    """Función principal para ejecutar el ETL."""
    # Configuración de directorios
    input_dir = "data/raw/guia"
    output_dir = "data/intermedio/procesados"
    errors_dir = "data/intermedio/errores"

    # Crear instancia y ejecutar
    etl = GuiaToIntermediateETL(input_dir, output_dir, errors_dir)
    etl.procesar_guia_urbana()


if __name__ == "__main__":
    main()