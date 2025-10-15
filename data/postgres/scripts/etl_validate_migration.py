#!/usr/bin/env python3
"""
ETL Migration Validation - Sprint 1 PostgreSQL Migration
======================================================

Este script valida completamente la migración de datos desde archivos
intermedios hasta PostgreSQL, comparando conteos, integridad y
realizando pruebas de rendimiento.

Flujo: Validación Comparativa → Pruebas Espaciales → Reporte Final

Author: Claude Code for Citrino
Date: October 2025
"""

import os
import json
import logging
import psycopg2
import psycopg2.extras
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/postgres/logs/etl_validate_migration.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class MigrationValidator:
    """Validador completo de la migración a PostgreSQL."""

    def __init__(self, intermediate_dir: str, output_dir: str):
        """
        Inicializar el validador.

        Args:
            intermediate_dir: Directorio con archivos intermedios
            output_dir: Directorio para reportes de validación
        """
        self.intermediate_dir = Path(intermediate_dir)
        self.output_dir = Path(output_dir)

        # Asegurar que el directorio de salida exista
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Configuración de base de datos
        self.db_config = {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME', 'citrino'),
            'user': os.getenv('DB_USER', 'postgres'),
            'password': os.getenv('DB_PASSWORD', '')
        }

        # Resultados de validación
        self.validation_results = {
            'conteos': {},
            'integridad': {},
            'calidad_datos': {},
            'rendimiento': {},
            'pruebas_espaciales': {},
            'errores': []
        }

        # Conexión a base de datos
        self.connection = None

    def conectar_db(self) -> bool:
        """
        Establecer conexión a la base de datos PostgreSQL.

        Returns:
            True si la conexión es exitosa
        """
        try:
            self.connection = psycopg2.connect(**self.db_config)
            logger.info(f"Conectado a PostgreSQL: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")
            return True
        except Exception as e:
            logger.error(f"Error conectando a PostgreSQL: {e}")
            return False

    def desconectar_db(self):
        """Cerrar conexión a la base de datos."""
        if self.connection:
            self.connection.close()
            logger.info("Conexión a PostgreSQL cerrada")

    def ejecutar_query(self, query: str, params: Optional[Tuple] = None) -> List[Tuple]:
        """
        Ejecutar query en la base de datos.

        Args:
            query: Query SQL a ejecutar
            params: Parámetros del query

        Returns:
            Resultados del query
        """
        try:
            with self.connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                cursor.execute(query, params)
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"Error ejecutando query: {e}")
            return []

    def contar_registros_excel(self) -> Dict[str, int]:
        """
        Contar registros en archivos Excel intermedios.

        Returns:
            Diccionario con conteos por tipo de archivo
        """
        conteos = {
            'agentes': 0,
            'propiedades': 0,
            'servicios': 0
        }

        try:
            # Contar agentes consolidados
            archivo_agentes = self.intermediate_dir / "agentes_consolidados.xlsx"
            if archivo_agentes.exists():
                df_agentes = pd.read_excel(archivo_agentes, sheet_name="Agentes Consolidados", engine='openpyxl')
                conteos['agentes'] = len(df_agentes)
                logger.info(f"Agentes en Excel: {conteos['agentes']}")

            # Contar propiedades (todos los archivos procesados)
            archivos_propiedades = list(self.intermediate_dir.glob("propiedades_*_procesado.xlsx"))
            for archivo in archivos_propiedades:
                try:
                    df_props = pd.read_excel(archivo, sheet_name="Propiedades", engine='openpyxl')
                    conteos['propiedades'] += len(df_props)
                except Exception as e:
                    logger.warning(f"Error leyendo {archivo.name}: {e}")

            logger.info(f"Propiedades en Excel: {conteos['propiedades']}")

            # Contar servicios
            archivos_servicios = list(self.intermediate_dir.glob("servicios_*_procesado.xlsx"))
            for archivo in archivos_servicios:
                try:
                    wb = pd.ExcelFile(archivo, engine='openpyxl')
                    total_servicios = 0
                    for sheet_name in wb.sheet_names:
                        if sheet_name not in ['Errores', 'Estadísticas', 'Metadatos']:
                            df_servicios = pd.read_excel(archivo, sheet_name=sheet_name, engine='openpyxl')
                            total_servicios += len(df_servicios)
                    conteos['servicios'] += total_servicios
                except Exception as e:
                    logger.warning(f"Error leyendo {archivo.name}: {e}")

            logger.info(f"Servicios en Excel: {conteos['servicios']}")

        except Exception as e:
            logger.error(f"Error contando registros en Excel: {e}")

        return conteos

    def contar_registros_postgres(self) -> Dict[str, int]:
        """
        Contar registros en PostgreSQL.

        Returns:
            Diccionario con conteos por tabla
        """
        conteos = {}

        try:
            # Tabla de agentes
            query_agentes = "SELECT COUNT(*) as total FROM agentes"
            resultado = self.ejecutar_query(query_agentes)
            conteos['agentes'] = resultado[0]['total'] if resultado else 0
            logger.info(f"Agentes en PostgreSQL: {conteos['agentes']}")

            # Tabla de propiedades
            query_propiedades = "SELECT COUNT(*) as total FROM propiedades"
            resultado = self.ejecutar_query(query_propiedades)
            conteos['propiedades'] = resultado[0]['total'] if resultado else 0
            logger.info(f"Propiedades en PostgreSQL: {conteos['propiedades']}")

            # Tabla de servicios
            query_servicios = "SELECT COUNT(*) as total FROM servicios"
            resultado = self.ejecutar_query(query_servicios)
            conteos['servicios'] = resultado[0]['total'] if resultado else 0
            logger.info(f"Servicios en PostgreSQL: {conteos['servicios']}")

        except Exception as e:
            logger.error(f"Error contando registros en PostgreSQL: {e}")

        return conteos

    def validar_conteos(self) -> Dict[str, Any]:
        """
        Validar que los conteos coincidan entre Excel y PostgreSQL.

        Returns:
            Resultados de validación de conteos
        """
        logger.info("Validando conteos de registros...")

        excel_conteos = self.contar_registros_excel()
        postgres_conteos = self.contar_registros_postgres()

        validacion = {
            'excel': excel_conteos,
            'postgres': postgres_conteos,
            'diferencias': {},
            'coincidencias': {},
            'porcentaje_migracion': {}
        }

        total_diferencias = 0

        for tabla in excel_conteos.keys():
            excel_count = excel_conteos[tabla]
            postgres_count = postgres_conteos.get(tabla, 0)

            diferencia = excel_count - postgres_count
            porcentaje = (postgres_count / excel_count * 100) if excel_count > 0 else 0

            validacion['diferencias'][tabla] = diferencia
            validacion['coincidencias'][tabla] = diferencia == 0
            validacion['porcentaje_migracion'][tabla] = round(porcentaje, 2)

            if diferencia != 0:
                total_diferencias += abs(diferencia)
                logger.warning(f"Diferencia en {tabla}: Excel={excel_count}, PostgreSQL={postgres_count}, Diferencia={diferencia}")

        validacion['total_diferencias'] = total_diferencias
        validacion['migracion_exitosa'] = total_diferencias == 0

        logger.info(f"Validación de conteos completada. Diferencias totales: {total_diferencias}")

        return validacion

    def validar_calidad_datos(self) -> Dict[str, Any]:
        """
        Validar calidad de datos en PostgreSQL.

        Returns:
            Resultados de validación de calidad
        """
        logger.info("Validando calidad de datos en PostgreSQL...")

        calidad = {}

        try:
            # Calidad de coordenadas en propiedades
            query_coords_props = """
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE coordenadas_validas = true) as validas,
                COUNT(*) FILTER (WHERE coordenadas IS NOT NULL) as con_coordenadas,
                COUNT(*) FILTER (WHERE datos_completos = true) as completos
            FROM propiedades
            """
            resultado = self.ejecutar_query(query_coords_props)
            if resultado:
                props_stats = resultado[0]
                calidad['propiedades'] = {
                    'total': props_stats['total'],
                    'coordenadas_validas': props_stats['validas'],
                    'con_coordenadas': props_stats['con_coordenadas'],
                    'datos_completos': props_stats['completos'],
                    'porcentaje_coords_validas': round(props_stats['validas'] / props_stats['total'] * 100, 2) if props_stats['total'] > 0 else 0,
                    'porcentaje_datos_completos': round(props_stats['completos'] / props_stats['total'] * 100, 2) if props_stats['total'] > 0 else 0
                }

            # Calidad de coordenadas en servicios
            query_coords_serv = """
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE coordenadas_validas = true) as validas,
                COUNT(*) FILTER (WHERE coordenadas IS NOT NULL) as con_coordenadas
            FROM servicios
            """
            resultado = self.ejecutar_query(query_coords_serv)
            if resultado:
                serv_stats = resultado[0]
                calidad['servicios'] = {
                    'total': serv_stats['total'],
                    'coordenadas_validas': serv_stats['validas'],
                    'con_coordenadas': serv_stats['con_coordenadas'],
                    'porcentaje_coords_validas': round(serv_stats['validas'] / serv_stats['total'] * 100, 2) if serv_stats['total'] > 0 else 0
                }

            # Calidad de datos de agentes
            query_agentes = """
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE email IS NOT NULL AND email != '') as con_email,
                COUNT(*) FILTER (WHERE telefono IS NOT NULL AND telefono != '') as con_telefono,
                COUNT(*) FILTER (WHERE empresa IS NOT NULL AND empresa != '') as con_empresa
            FROM agentes
            """
            resultado = self.ejecutar_query(query_agentes)
            if resultado:
                ag_stats = resultado[0]
                calidad['agentes'] = {
                    'total': ag_stats['total'],
                    'con_email': ag_stats['con_email'],
                    'con_telefono': ag_stats['con_telefono'],
                    'con_empresa': ag_stats['con_empresa'],
                    'porcentaje_con_email': round(ag_stats['con_email'] / ag_stats['total'] * 100, 2) if ag_stats['total'] > 0 else 0,
                    'porcentaje_con_telefono': round(ag_stats['con_telefono'] / ag_stats['total'] * 100, 2) if ag_stats['total'] > 0 else 0
                }

            # Distribución por zonas
            query_zonas = """
            SELECT zona, COUNT(*) as total_propiedades, AVG(precio_usd) as precio_promedio
            FROM propiedades
            WHERE zona IS NOT NULL AND zona != ''
            GROUP BY zona
            ORDER BY total_propiedades DESC
            LIMIT 10
            """
            resultado = self.ejecutar_query(query_zonas)
            calidad['top_zonas'] = [dict(row) for row in resultado]

        except Exception as e:
            logger.error(f"Error validando calidad de datos: {e}")
            calidad['error'] = str(e)

        return calidad

    def probar_rendimiento_espacial(self) -> Dict[str, Any]:
        """
        Probar rendimiento de consultas espaciales.

        Returns:
            Resultados de pruebas de rendimiento
        """
        logger.info("Probando rendimiento de consultas espaciales...")

        rendimiento = {}

        try:
            # Prueba 1: Búsqueda de propiedades cerca de servicios
            start_time = time.time()
            query_cercanos = """
            SELECT COUNT(*) as total
            FROM propiedades p
            JOIN servicios s ON ST_DWithin(p.coordenadas, s.coordenadas, 500)
            WHERE p.coordenadas_validas = true AND s.coordenadas_validas = true
            """
            resultado = self.ejecutar_query(query_cercanos)
            tiempo_cercanos = time.time() - start_time

            rendimiento['busqueda_cercanos'] = {
                'resultado': resultado[0]['total'] if resultado else 0,
                'tiempo_segundos': round(tiempo_cercanos, 3),
                'rendimiento': 'Excelente' if tiempo_cercanos < 1 else 'Bueno' if tiempo_cercanos < 5 else 'Regular'
            }

            # Prueba 2: Cálculo de distancias
            start_time = time.time()
            query_distancias = """
            SELECT
                p.titulo,
                s.nombre as servicio_nombre,
                ST_Distance(p.coordenadas, s.coordenadas) as distancia
            FROM propiedades p
            CROSS JOIN servicios s
            WHERE p.coordenadas_validas = true
            AND s.coordenadas_validas = true
            AND s.tipo_servicio = 'Educación'
            LIMIT 100
            """
            resultado = self.ejecutar_query(query_distancias)
            tiempo_distancias = time.time() - start_time

            rendimiento['calculo_distancias'] = {
                'resultados': len(resultado),
                'tiempo_segundos': round(tiempo_distancias, 3),
                'rendimiento': 'Excelente' if tiempo_distancias < 2 else 'Bueno' if tiempo_distancias < 10 else 'Regular'
            }

            # Prueba 3: Análisis por zona
            start_time = time.time()
            query_zonas = """
            SELECT
                p.zona,
                COUNT(*) as total_propiedades,
                COUNT(DISTINCT s.id) as servicios_cercanos
            FROM propiedades p
            LEFT JOIN servicios s ON ST_DWithin(p.coordenadas, s.coordenadas, 1000)
            WHERE p.zona IS NOT NULL
            GROUP BY p.zona
            ORDER BY total_propiedades DESC
            """
            resultado = self.ejecutar_query(query_zonas)
            tiempo_zonas = time.time() - start_time

            rendimiento['analisis_zonas'] = {
                'zonas_procesadas': len(resultado),
                'tiempo_segundos': round(tiempo_zonas, 3),
                'rendimiento': 'Excelente' if tiempo_zonas < 3 else 'Bueno' if tiempo_zonas < 15 else 'Regular'
            }

            # Prueba 4: Índices espaciales
            start_time = time.time()
            query_indices = """
            SELECT indexname, tablename
            FROM pg_indexes
            WHERE tablename IN ('propiedades', 'servicios')
            AND indexname LIKE '%coordenadas%'
            """
            resultado = self.ejecutar_query(query_indices)
            rendimiento['indices_espaciales'] = {
                'cantidad': len(resultado),
                'indices': [dict(row) for row in resultado]
            }

        except Exception as e:
            logger.error(f"Error en pruebas de rendimiento: {e}")
            rendimiento['error'] = str(e)

        return rendimiento

    def validar_integridad_referencial(self) -> Dict[str, Any]:
        """
        Validar integridad referencial de los datos.

        Returns:
            Resultados de validación de integridad
        """
        logger.info("Validando integridad referencial...")

        integridad = {}

        try:
            # Verificar relaciones entre propiedades y agentes (si existen)
            query_props_agentes = """
            SELECT
                COUNT(*) as total_props,
                COUNT(*) FILTER (WHERE agente_id IS NOT NULL) as con_agente,
                COUNT(DISTINCT agente_id) as agentes_unicos
            FROM propiedades
            """
            resultado = self.ejecutar_query(query_props_agentes)
            if resultado:
                props_stats = resultado[0]
                integridad['propiedades_agentes'] = {
                    'total_propiedades': props_stats['total_props'],
                    'con_agente_asignado': props_stats['con_agente'],
                    'agentes_unicos_referenciados': props_stats['agentes_unicos'],
                    'porcentaje_con_agente': round(props_stats['con_agente'] / props_stats['total_props'] * 100, 2) if props_stats['total_props'] > 0 else 0
                }

            # Verificar unicidad de emails en agentes
            query_emails_unicos = """
            SELECT email, COUNT(*) as count
            FROM agentes
            WHERE email IS NOT NULL AND email != ''
            GROUP BY email
            HAVING COUNT(*) > 1
            """
            resultado = self.ejecutar_query(query_emails_unicos)
            integridad['emails_duplicados'] = {
                'cantidad': len(resultado),
                'detalles': [dict(row) for row in resultado[:5]]  # Primeros 5 duplicados
            }

            # Verificar coordenadas fuera de bounds
            query_fuera_bounds = """
            SELECT COUNT(*) as total
            FROM propiedades
            WHERE coordenadas_validas = false
            """
            resultado = self.ejecutar_query(query_fuera_bounds)
            integridad['coordenadas_fuera_bounds'] = {
                'propiedades': resultado[0]['total'] if resultado else 0
            }

            query_servicios_fuera_bounds = """
            SELECT COUNT(*) as total
            FROM servicios
            WHERE coordenadas_validas = false
            """
            resultado = self.ejecutar_query(query_servicios_fuera_bounds)
            integridad['coordenadas_fuera_bounds']['servicios'] = resultado[0]['total'] if resultado else 0

        except Exception as e:
            logger.error(f"Error validando integridad referencial: {e}")
            integridad['error'] = str(e)

        return integridad

    def generar_reporte_excel(self) -> str:
        """
        Generar reporte completo en Excel.

        Returns:
            Ruta del archivo generado
        """
        logger.info("Generando reporte de validación en Excel...")

        output_filename = f"reporte_validacion_migracion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = self.output_dir / output_filename

        wb = Workbook()
        wb.remove(wb.active)

        # 1. Hoja de Resumen Ejecutivo
        ws_resumen = wb.create_sheet("Resumen Ejecutivo")

        resumen_data = [
            ['Métrica', 'Valor', 'Estado'],
            ['Fecha Validación', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
            ['', '', ''],
            ['Agentes', f"{self.validation_results['conteos']['postgres'].get('agentes', 0)} cargados",
             '✓' if self.validation_results['conteos']['coincidencias'].get('agentes', False) else '✗'],
            ['Propiedades', f"{self.validation_results['conteos']['postgres'].get('propiedades', 0)} cargadas",
             '✓' if self.validation_results['conteos']['coincidencias'].get('propiedades', False) else '✗'],
            ['Servicios', f"{self.validation_results['conteos']['postgres'].get('servicios', 0)} cargados",
             '✓' if self.validation_results['conteos']['coincidencias'].get('servicios', False) else '✗'],
            ['', '', ''],
            ['Migración General',
             'Exitosa' if self.validation_results['conteos']['migracion_exitosa'] else 'Con Diferencias',
             '✓' if self.validation_results['conteos']['migracion_exitosa'] else '⚠']
        ]

        for row in resumen_data:
            ws_resumen.append(row)

        # Formato de encabezados
        for cell in ws_resumen[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # 2. Hoja de Comparación de Conteos
        ws_conteos = wb.create_sheet("Comparación de Conteos")

        # Datos de comparación
        conteos_data = [['Tabla', 'Excel', 'PostgreSQL', 'Diferencia', '% Migración', 'Estado']]
        for tabla in ['agentes', 'propiedades', 'servicios']:
            excel_count = self.validation_results['conteos']['excel'].get(tabla, 0)
            pg_count = self.validation_results['conteos']['postgres'].get(tabla, 0)
            diferencia = excel_count - pg_count
            porcentaje = self.validation_results['conteos']['porcentaje_migracion'].get(tabla, 0)
            estado = '✓ OK' if diferencia == 0 else f'⚠ Dif: {diferencia}'

            conteos_data.append([tabla.title(), excel_count, pg_count, diferencia, f"{porcentaje}%", estado])

        for row in conteos_data:
            ws_conteos.append(row)

        # Formato y gráfico
        for cell in ws_conteos[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 3. Hoja de Calidad de Datos
        ws_calidad = wb.create_sheet("Calidad de Datos")

        calidad_data = [['Métrica', 'Total', 'Válidos', 'Porcentaje', 'Estado']]

        # Propiedades
        if 'propiedades' in self.validation_results['calidad_datos']:
            props = self.validation_results['calidad_datos']['propiedades']
            calidad_data.extend([
                ['Propiedades - Coordenadas Válidas', props['total'], props['coordenadas_validas'],
                 f"{props['porcentaje_coords_validas']}%",
                 '✓' if props['porcentaje_coords_validas'] > 80 else '⚠'],
                ['Propiedades - Datos Completos', props['total'], props['datos_completos'],
                 f"{props['porcentaje_datos_completos']}%",
                 '✓' if props['porcentaje_datos_completos'] > 80 else '⚠']
            ])

        # Servicios
        if 'servicios' in self.validation_results['calidad_datos']:
            serv = self.validation_results['calidad_datos']['servicios']
            calidad_data.extend([
                ['Servicios - Coordenadas Válidas', serv['total'], serv['coordenadas_validas'],
                 f"{serv['porcentaje_coords_validas']}%",
                 '✓' if serv['porcentaje_coords_validas'] > 80 else '⚠']
            ])

        # Agentes
        if 'agentes' in self.validation_results['calidad_datos']:
            agentes = self.validation_results['calidad_datos']['agentes']
            calidad_data.extend([
                ['Agentes - Con Email', agentes['total'], agentes['con_email'],
                 f"{agentes['porcentaje_con_email']}%",
                 '✓' if agentes['porcentaje_con_email'] > 60 else '⚠'],
                ['Agentes - Con Teléfono', agentes['total'], agentes['con_telefono'],
                 f"{agentes['porcentaje_con_telefono']}%",
                 '✓' if agentes['porcentaje_con_telefono'] > 60 else '⚠']
            ])

        for row in calidad_data:
            ws_calidad.append(row)

        # Formato
        for cell in ws_calidad[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 4. Hoja de Rendimiento
        ws_rendimiento = wb.create_sheet("Rendimiento")

        rendimiento_data = [['Prueba', 'Resultado', 'Tiempo (s)', 'Estado']]
        for prueba, datos in self.validation_results['rendimiento'].items():
            if isinstance(datos, dict) and 'tiempo_segundos' in datos:
                estado = datos.get('rendimiento', 'Desconocido')
                icono = '✓' if estado == 'Excelente' else '✓' if estado == 'Bueno' else '⚠'
                rendimiento_data.append([
                    prueba.replace('_', ' ').title(),
                    datos.get('resultado', datos.get('resultados', 'N/A')),
                    datos['tiempo_segundos'],
                    f"{icono} {estado}"
                ])

        for row in rendimiento_data:
            ws_rendimiento.append(row)

        # Formato
        for cell in ws_rendimiento[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 5. Hoja de Detalles y Errores
        ws_detalles = wb.create_sheet("Detalles y Errores")

        detalles_data = [['Categoría', 'Detalle', 'Gravedad']]

        # Agregar errores o advertencias
        if self.validation_results['conteos']['total_diferencias'] > 0:
            detalles_data.append([
                'Conteos',
                f"Diferencias totales: {self.validation_results['conteos']['total_diferencias']}",
                'Media'
            ])

        if 'integridad' in self.validation_results:
            integridad = self.validation_results['integridad']
            if 'emails_duplicados' in integridad and integridad['emails_duplicados']['cantidad'] > 0:
                detalles_data.append([
                    'Integridad',
                    f"Emails duplicados: {integridad['emails_duplicados']['cantidad']}",
                    'Alta'
                ])

        if len(detalles_data) == 1:  # Solo encabezado
            detalles_data.append(['Estado', 'No se detectaron problemas graves', '✓'])

        for row in detalles_data:
            ws_detalles.append(row)

        # Formato
        for cell in ws_detalles[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Reporte de validación guardado: {output_path}")

        return str(output_path)

    def ejecutar_validacion_completa(self):
        """
        Ejecutar validación completa de la migración.
        """
        logger.info("Iniciando validación completa de la migración PostgreSQL...")

        if not self.conectar_db():
            logger.error("No se pudo conectar a la base de datos para validación")
            return

        try:
            # 1. Validar conteos
            self.validation_results['conteos'] = self.validar_conteos()

            # 2. Validar calidad de datos
            self.validation_results['calidad_datos'] = self.validar_calidad_datos()

            # 3. Validar integridad referencial
            self.validation_results['integridad'] = self.validar_integridad_referencial()

            # 4. Probar rendimiento espacial
            self.validation_results['rendimiento'] = self.probar_rendimiento_espacial()

            # 5. Generar reporte Excel
            reporte_path = self.generar_reporte_excel()

            # 6. Generar reporte JSON
            reporte_json_path = self.output_dir / f"reporte_validacion_json_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(reporte_json_path, 'w', encoding='utf-8') as f:
                json.dump(self.validation_results, f, indent=2, ensure_ascii=False, default=str)

            # 7. Imprimir resumen final
            self.imprimir_resumen_final()

            logger.info("Validación completa finalizada exitosamente")
            logger.info(f"Reporte Excel: {reporte_path}")
            logger.info(f"Reporte JSON: {reporte_json_path}")

        except Exception as e:
            logger.error(f"Error en validación completa: {e}")
        finally:
            self.desconectar_db()

    def imprimir_resumen_final(self):
        """Imprimir resumen final de la validación."""
        logger.info("=" * 80)
        logger.info("RESUMEN FINAL DE VALIDACIÓN - MIGRACIÓN POSTGRESQL")
        logger.info("=" * 80)

        # Conteos
        conteos = self.validation_results['conteos']
        logger.info(f"ESTADO MIGRACIÓN: {'✓ EXITOSA' if conteos['migracion_exitosa'] else '⚠ CON DIFERENCIAS'}")
        logger.info("")
        logger.info("Conteos de registros:")
        for tabla in ['agentes', 'propiedades', 'servicios']:
            excel_count = conteos['excel'].get(tabla, 0)
            pg_count = conteos['postgres'].get(tabla, 0)
            porc = conteos['porcentaje_migracion'].get(tabla, 0)
            estado = "✓" if conteos['coincidencias'].get(tabla, False) else f"✗ (Dif: {excel_count - pg_count})"
            logger.info(f"  {tabla.title()}: Excel={excel_count}, PostgreSQL={pg_count}, {porc}% {estado}")

        logger.info(f"Diferencias totales: {conteos['total_diferencias']}")

        # Calidad de datos
        if 'calidad_datos' in self.validation_results:
            calidad = self.validation_results['calidad_datos']
            logger.info("")
            logger.info("Calidad de datos:")
            if 'propiedades' in calidad:
                props = calidad['propiedades']
                logger.info(f"  Propiedades con coordenadas válidas: {props['coordenadas_validas']}/{props['total']} ({props['porcentaje_coords_validas']}%)")
            if 'servicios' in calidad:
                serv = calidad['servicios']
                logger.info(f"  Servicios con coordenadas válidas: {serv['coordenadas_validas']}/{serv['total']} ({serv['porcentaje_coords_validas']}%)")
            if 'agentes' in calidad:
                agentes = calidad['agentes']
                logger.info(f"  Agentes con email: {agentes['con_email']}/{agentes['total']} ({agentes['porcentaje_con_email']}%)")

        # Rendimiento
        if 'rendimiento' in self.validation_results:
            rendimiento = self.validation_results['rendimiento']
            logger.info("")
            logger.info("Rendimiento de consultas espaciales:")
            for prueba, datos in rendimiento.items():
                if isinstance(datos, dict) and 'tiempo_segundos' in datos:
                    estado = datos.get('rendimiento', 'Desconocido')
                    logger.info(f"  {prueba.replace('_', ' ').title()}: {datos['tiempo_segundos']}s ({estado})")

        # Problemas detectados
        problemas = []
        if conteos['total_diferencias'] > 0:
            problemas.append(f"Diferencias en conteos: {conteos['total_diferencias']}")

        if 'integridad' in self.validation_results:
            integridad = self.validation_results['integridad']
            if 'emails_duplicados' in integridad and integridad['emails_duplicados']['cantidad'] > 0:
                problemas.append(f"Emails duplicados: {integridad['emails_duplicados']['cantidad']}")

        if problemas:
            logger.info("")
            logger.warning("Problemas detectados:")
            for problema in problemas:
                logger.warning(f"  ⚠ {probleblema}")
        else:
            logger.info("")
            logger.info("✓ No se detectaron problemas graves")

        logger.info("=" * 80)


def main():
    """Función principal para ejecutar la validación."""
    # Configuración de directorios
    intermediate_dir = "data/intermedio/procesados"
    output_dir = "data/postgres/logs"

    # Verificar configuración de base de datos
    if not os.getenv('DB_HOST'):
        logger.error("Variables de entorno de base de datos no configuradas")
        return

    # Crear instancia y ejecutar
    validator = MigrationValidator(intermediate_dir, output_dir)
    validator.ejecutar_validacion_completa()


if __name__ == "__main__":
    main()