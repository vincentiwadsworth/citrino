#!/usr/bin/env python3
"""
ETL Consolidar Agentes - Sprint 1 PostgreSQL Migration
=====================================================

Este script consolida y deduplica agentes inmobiliarios de múltiples
archivos intermedios de propiedades, generando un maestro único.

Flujo: Múltiples Agentes → Deduplicación → Consolidación → Maestro Único

Author: Claude Code for Citrino
Date: October 2025
"""

import os
import re
import json
import logging
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from difflib import SequenceMatcher
import hashlib

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/postgres/logs/etl_consolidar_agentes.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class AgentConsolidationETL:
    """ETL para consolidar y deduplicar agentes inmobiliarios."""

    def __init__(self, input_dir: str, output_dir: str, errors_dir: str):
        """
        Inicializar el ETL.

        Args:
            input_dir: Directorio con archivos procesados
            output_dir: Directorio para archivo consolidado
            errors_dir: Directorio para logs de errores
        """
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.errors_dir = Path(errors_dir)

        # Asegurar que los directorios existan
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.errors_dir.mkdir(parents=True, exist_ok=True)

        # Estadísticas del proceso
        self.stats = {
            'total_agentes_raw': 0,
            'agentes_unicos': 0,
            'duplicados_eliminados': 0,
            'agentes_con_email': 0,
            'agentes_con_telefono': 0,
            'agentes_con_empresa': 0,
            'grupos_duplicados': 0
        }

        # Patrones comunes de nombres y empresas
        self.nombres_comunes = {
            'john', 'jane', 'maria', 'jose', 'carlos', 'luis', 'ana', 'martha',
            'juan', 'pedro', 'pablo', 'david', 'mario', 'laura', 'claudia',
            'agente', 'asesor', 'corredor', 'broker', 'inmobiliaria'
        }

        self.empresas_conocidas = {
            'citrino', 'remax', 'coldwell', 'cb', 'century21', 'propiedades',
            'inmobiliaria', 'inmuebles', 'bienes', 'raices', 'real state',
            'garcia', 'rojas', 'fernandez', 'perez', 'gonzalez', 'rivera'
        }

    def limpiar_texto(self, texto: str) -> str:
        """
        Limpiar y normalizar texto.

        Args:
            texto: Texto a limpiar

        Returns:
            Texto limpio y normalizado
        """
        if pd.isna(texto) or texto is None:
            return ""

        # Convertir a string y normalizar Unicode
        texto = str(texto)
        texto = unicodedata.normalize('NFKD', texto)

        # Quitar caracteres extraños y espacios múltiples
        texto = re.sub(r'[^\w\s.-]', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto)

        return texto.strip().title()

    def normalizar_telefono(self, telefono: str) -> Optional[str]:
        """
        Normalizar número de teléfono boliviano.

        Args:
            telefono: Teléfono en cualquier formato

        Returns:
            Teléfono normalizado o None
        """
        if pd.isna(telefono) or telefono is None:
            return None

        # Extraer solo dígitos
        digitos = re.findall(r'\d', str(telefono))

        if len(digitos) == 8:
            # Formato: 1234-5678
            return f"{''.join(digitos[:4])}-{''.join(digitos[4:])}"
        elif len(digitos) == 7:
            # Formato antiguo: 1234567
            return ''.join(digitos)
        else:
            return None

    def validar_email(self, email: str) -> Optional[str]:
        """
        Validar y normalizar email.

        Args:
            email: Email a validar

        Returns:
            Email normalizado o None si es inválido
        """
        if pd.isna(email) or email is None:
            return None

        email = str(email).lower().strip()

        # Patrón básico de email
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'

        if re.match(email_pattern, email):
            return email

        return None

    def generar_hash_agente(self, nombre: str, email: str = "", telefono: str = "") -> str:
        """
        Generar hash único para identificar un agente.

        Args:
            nombre: Nombre del agente
            email: Email del agente
            telefono: Teléfono del agente

        Returns:
            Hash SHA256 del agente
        """
        # Crear string de identificación
        identificacion = f"{nombre.lower()}|{email.lower()}|{telefono}"
        return hashlib.sha256(identificacion.encode()).hexdigest()[:16]

    def similitud_nombres(self, nombre1: str, nombre2: str) -> float:
        """
        Calcular similitud entre dos nombres.

        Args:
            nombre1: Primer nombre
            nombre2: Segundo nombre

        Returns:
            Puntaje de similitud (0-1)
        """
        return SequenceMatcher(None, nombre1.lower(), nombre2.lower()).ratio()

    def son_mismo_agente(self, agente1: Dict, agente2: Dict) -> bool:
        """
        Determinar si dos registros representan al mismo agente.

        Args:
            agente1: Primer registro de agente
            agente2: Segundo registro de agente

        Returns:
            True si son el mismo agente
        """
        # 1. Email exacto
        if agente1['email'] and agente2['email']:
            if agente1['email'] == agente2['email']:
                return True

        # 2. Teléfono exacto
        if agente1['telefono'] and agente2['telefono']:
            if agente1['telefono'] == agente2['telefono']:
                return True

        # 3. Similitud de nombres + mismo email/telefono
        if agente1['nombre'] and agente2['nombre']:
            similitud = self.similitud_nombres(agente1['nombre'], agente2['nombre'])

            # Si los nombres son muy similares y comparten email o teléfono
            if similitud > 0.8:
                if (agente1['email'] and agente2['email']) or (agente1['telefono'] and agente2['telefono']):
                    return True

            # Si los nombres son idénticos
            if similitud > 0.95:
                return True

        # 4. Nombre similar + misma empresa
        if (agente1['nombre'] and agente2['nombre'] and
            agente1['empresa'] and agente2['empresa']):
            if (self.similitud_nombres(agente1['nombre'], agente2['nombre']) > 0.7 and
                agente1['empresa'] == agente2['empresa']):
                return True

        return False

    def consolidar_informacion_agente(self, agentes_grupo: List[Dict]) -> Dict:
        """
        Consolidar información de múltiples registros del mismo agente.

        Args:
            agentes_grupo: Lista de registros del mismo agente

        Returns:
            Agente consolidado
        """
        # Prioridad: el registro más completo
        agentes_prioridad = sorted(agentes_grupo, key=lambda x: sum([
            1 if x['nombre'] and len(x['nombre']) > 5 else 0,
            1 if x['email'] else 0,
            1 if x['telefono'] else 0,
            1 if x['empresa'] and len(x['empresa']) > 3 else 0
        ]), reverse=True)

        agente_base = agentes_prioridad[0].copy()

        # Complementar información de otros registros
        for agente in agentes_grupo[1:]:
            if not agente_base['email'] and agente['email']:
                agente_base['email'] = agente['email']

            if not agente_base['telefono'] and agente['telefono']:
                agente_base['telefono'] = agente['telefono']

            if not agente_base['empresa'] and agente['empresa']:
                agente_base['empresa'] = agente['empresa']

            # Conservar fuente del más completo
            if len(agente['fuente_archivo']) > len(agente_base['fuente_archivo']):
                agente_base['fuente_archivo'] = agente['fuente_archivo']

        # Limpiar y normalizar campos finales
        agente_base['nombre'] = self.limpiar_texto(agente_base['nombre'])
        agente_base['email'] = self.validar_email(agente_base['email'])
        agente_base['telefono'] = self.normalizar_telefono(agente_base['telefono'])
        agente_base['empresa'] = self.limpiar_texto(agente_base['empresa'])

        # Generar hash único
        agente_base['hash_agente'] = self.generar_hash_agente(
            agente_base['nombre'],
            agente_base['email'] or "",
            agente_base['telefono'] or ""
        )

        # Metadatos
        agente_base['registros_consolidados'] = len(agentes_grupo)
        agente_base['fuentes_originales'] = list(set([a['fuente_archivo'] for a in agentes_grupo]))
        agente_base['fecha_consolidacion'] = datetime.now()

        return agente_base

    def procesar_agentes_de_archivos(self) -> Dict[str, Any]:
        """
        Procesar todos los archivos intermedios para extraer agentes.

        Returns:
            Diccionario con resultados del procesamiento
        """
        logger.info("Procesando agentes de archivos intermedios...")

        todos_agentes = []
        archivos_procesados = []

        # Buscar archivos Excel procesados
        archivos_excel = list(self.input_dir.glob("propiedades_*_procesado.xlsx"))

        for archivo_path in archivos_excel:
            try:
                logger.info(f"Leyendo agentes de: {archivo_path.name}")

                # Leer Excel
                df = pd.read_excel(archivo_path, sheet_name="Agentes", engine='openpyxl')

                # Estandarizar nombres de columnas
                df.columns = [col.lower().replace(' ', '_') for col in df.columns]

                # Extraer agentes
                for idx, row in df.iterrows():
                    if pd.isna(row.get('nombre')) or not row.get('nombre'):
                        continue

                    agente = {
                        'nombre': self.limpiar_texto(row.get('nombre', '')),
                        'telefono': self.normalizar_telefono(row.get('telefono', '')),
                        'email': self.validar_email(row.get('email', '')),
                        'empresa': self.limpiar_texto(row.get('empresa', '')),
                        'fuente_archivo': archivo_path.name,
                        'fila_original': idx + 1
                    }

                    todos_agentes.append(agente)

                archivos_procesados.append(archivo_path.name)
                logger.info(f"Extraídos {len(df)} agentes de {archivo_path.name}")

            except Exception as e:
                logger.error(f"Error procesando archivo {archivo_path}: {e}")
                continue

        self.stats['total_agentes_raw'] = len(todos_agentes)
        logger.info(f"Total agentes extraídos: {len(todos_agentes)}")

        return {
            'agentes': todos_agentes,
            'archivos_procesados': archivos_procesados
        }

    def deduplicar_agentes(self, agentes: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
        """
        Deduplicar agentes usando similitud y reglas de negocio.

        Args:
            agentes: Lista de agentes a deduplicar

        Returns:
            Tupla (agentes_unicos, grupos_duplicados)
        """
        logger.info("Iniciando deduplicación de agentes...")

        agentes_unicos = []
        grupos_duplicados = []
        procesados = set()

        for i, agente_actual in enumerate(agentes):
            if i in procesados:
                continue

            # Buscar duplicados
            grupo_duplicados = [agente_actual]
            duplicados_indices = [i]

            for j, otro_agente in enumerate(agentes[i+1:], i+1):
                if j in procesados:
                    continue

                if self.son_mismo_agente(agente_actual, otro_agente):
                    grupo_duplicados.append(otro_agente)
                    duplicados_indices.append(j)

            # Marcar como procesados
            for idx in duplicados_indices:
                procesados.add(idx)

            # Consolidar información
            if len(grupo_duplicados) > 1:
                agente_consolidado = self.consolidar_informacion_agente(grupo_duplicados)
                grupos_duplicados.append({
                    'hash_agente': agente_consolidado['hash_agente'],
                    'registros_originales': len(grupo_duplicados),
                    'archivos': list(set([a['fuente_archivo'] for a in grupo_duplicados])),
                    'detalles': grupo_duplicados
                })
                self.stats['duplicados_eliminados'] += len(grupo_duplicados) - 1
                self.stats['grupos_duplicados'] += 1
            else:
                agente_consolidado = grupo_duplicados[0].copy()
                agente_consolidado['hash_agente'] = self.generar_hash_agente(
                    agente_consolidado['nombre'],
                    agente_consolidado['email'] or "",
                    agente_consolidado['telefono'] or ""
                )
                agente_consolidado['registros_consolidados'] = 1
                agente_consolidado['fuentes_originales'] = [agente_consolidado['fuente_archivo']]
                agente_consolidado['fecha_consolidacion'] = datetime.now()

            agentes_unicos.append(agente_consolidado)

        # Actualizar estadísticas
        self.stats['agentes_unicos'] = len(agentes_unicos)
        self.stats['agentes_con_email'] = len([a for a in agentes_unicos if a['email']])
        self.stats['agentes_con_telefono'] = len([a for a in agentes_unicos if a['telefono']])
        self.stats['agentes_con_empresa'] = len([a for a in agentes_unicos if a['empresa']])

        logger.info(f"Deduplicación completada:")
        logger.info(f"  - Agentes únicos: {len(agentes_unicos)}")
        logger.info(f"  - Duplicados eliminados: {self.stats['duplicados_eliminados']}")
        logger.info(f"  - Grupos de duplicados: {self.stats['grupos_duplicados']}")

        return agentes_unicos, grupos_duplicados

    def guardar_archivo_consolidado(self, agentes_unicos: List[Dict], grupos_duplicados: List[Dict], archivos_procesados: List[str]):
        """
        Guardar agentes consolidados en archivo Excel.

        Args:
            agentes_unicos: Lista de agentes únicos
            grupos_duplicados: Información de grupos duplicados
            archivos_procesados: Lista de archivos procesados
        """
        output_filename = "agentes_consolidados.xlsx"
        output_path = self.output_dir / output_filename

        # Crear workbook
        wb = Workbook()

        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # 1. Hoja principal de Agentes Consolidados
        ws_agentes = wb.create_sheet("Agentes Consolidados")

        # Preparar DataFrame de agentes únicos
        df_agentes = pd.DataFrame(agentes_unicos)

        # Reordenar columnas para mejor visualización
        columnas_orden = [
            'hash_agente', 'nombre', 'email', 'telefono', 'empresa',
            'registros_consolidados', 'fuentes_originales', 'fecha_consolidacion'
        ]
        df_agentes = df_agentes.reindex(columns=columnas_orden)

        # Escribir datos
        for r in dataframe_to_rows(df_agentes, index=False, header=True):
            ws_agentes.append(r)

        # Formato de encabezados
        for cell in ws_agentes[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Autoajustar ancho de columnas
        for column in ws_agentes.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_agentes.column_dimensions[column_letter].width = adjusted_width

        # 2. Hoja de Grupos Duplicados
        if grupos_duplicados:
            ws_duplicados = wb.create_sheet("Grupos Duplicados")

            # Preparar datos de duplicados
            datos_duplicados = []
            for grupo in grupos_duplicados:
                datos_duplicados.append({
                    'Hash Agente': grupo['hash_agente'],
                    'Registros Originales': grupo['registros_originales'],
                    'Archivos': ', '.join(grupo['archivos']),
                    'Ejemplo Nombre': grupo['detalles'][0]['nombre'],
                    'Ejemplo Email': grupo['detalles'][0]['email'],
                    'Ejemplo Teléfono': grupo['detalles'][0]['telefono']
                })

            df_duplicados = pd.DataFrame(datos_duplicados)

            for r in dataframe_to_rows(df_duplicados, index=False, header=True):
                ws_duplicados.append(r)

            # Formato de encabezados
            for cell in ws_duplicados[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # 3. Hoja de Estadísticas
        ws_stats = wb.create_sheet("Estadísticas")

        stats_data = [
            ['Métrica', 'Valor'],
            ['Total agentes procesados', self.stats['total_agentes_raw']],
            ['Agentes únicos', self.stats['agentes_unicos']],
            ['Duplicados eliminados', self.stats['duplicados_eliminados']],
            ['Grupos de duplicados', self.stats['grupos_duplicados']],
            ['Agentes con email', self.stats['agentes_con_email']],
            ['Agentes con teléfono', self.stats['agentes_con_telefono']],
            ['Agentes con empresa', self.stats['agentes_con_empresa']],
            ['Porcentaje con email', f"{round(self.stats['agentes_con_email']/self.stats['agentes_unicos']*100, 2)}%" if self.stats['agentes_unicos'] > 0 else "0%"],
            ['Porcentaje con teléfono', f"{round(self.stats['agentes_con_telefono']/self.stats['agentes_unicos']*100, 2)}%" if self.stats['agentes_unicos'] > 0 else "0%"],
            ['Archivos procesados', len(archivos_procesados)],
            ['Fecha consolidación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Versión script', '1.0.0']
        ]

        for row in stats_data:
            ws_stats.append(row)

        # Formato de encabezados
        for cell in ws_stats[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 4. Hoja de Metadatos
        ws_metadatos = wb.create_sheet("Metadatos")

        metadatos_data = [
            ['Propiedad', 'Valor'],
            ['Archivos procesados', ', '.join(archivos_procesados)],
            ['Fecha consolidación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Versión script', '1.0.0'],
            ['Algoritmo deduplicación', 'Similitud de nombres + Email/Teléfono exactos'],
            ['Umbral similitud nombres', '0.8 (con contacto) / 0.95 (sin contacto)'],
            ['Responsable', 'ETL Automático']
        ]

        for row in metadatos_data:
            ws_metadatos.append(row)

        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Archivo consolidado guardado: {output_path}")

        return output_path

    def generar_reporte_final(self, agentes_unicos: List[Dict], archivos_procesados: List[str]):
        """
        Generar reporte final del proceso de consolidación.

        Args:
            agentes_unicos: Lista de agentes únicos
            archivos_procesados: Archivos procesados
        """
        # Analizar calidad de datos
        calidad_datos = {
            'nombres_completos': len([a for a in agentes_unicos if a['nombre'] and len(a['nombre']) > 5]),
            'emails_validos': len([a for a in agentes_unicos if a['email']]),
            'telefonos_validos': len([a for a in agentes_unicos if a['telefono']]),
            'empresas_registradas': len([a for a in agentes_unicos if a['empresa']]),
            'multiples_registros': len([a for a in agentes_unicos if a['registros_consolidados'] > 1])
        }

        # Top empresas
        empresas_count = {}
        for agente in agentes_unicos:
            if agente['empresa']:
                empresas_count[agente['empresa']] = empresas_count.get(agente['empresa'], 0) + 1

        top_empresas = sorted(empresas_count.items(), key=lambda x: x[1], reverse=True)[:10]

        reporte = {
            'resumen_consolidacion': {
                'fecha': datetime.now().isoformat(),
                'total_agentes_raw': self.stats['total_agentes_raw'],
                'agentes_unicos': self.stats['agentes_unicos'],
                'duplicados_eliminados': self.stats['duplicados_eliminados'],
                'grupos_duplicados': self.stats['grupos_duplicados'],
                'archivos_procesados': len(archivos_procesados)
            },
            'calidad_datos': calidad_datos,
            'metricas_calidad': {
                'porcentaje_nombres_completos': round(calidad_datos['nombres_completos'] / self.stats['agentes_unicos'] * 100, 2) if self.stats['agentes_unicos'] > 0 else 0,
                'porcentaje_emails_validos': round(calidad_datos['emails_validos'] / self.stats['agentes_unicos'] * 100, 2) if self.stats['agentes_unicos'] > 0 else 0,
                'porcentaje_telefonos_validos': round(calidad_datos['telefonos_validos'] / self.stats['agentes_unicos'] * 100, 2) if self.stats['agentes_unicos'] > 0 else 0,
                'tasa_duplicacion': round(self.stats['duplicados_eliminados'] / self.stats['total_agentes_raw'] * 100, 2) if self.stats['total_agentes_raw'] > 0 else 0
            },
            'top_empresas': dict(top_empresas),
            'archivos_procesados': archivos_procesados
        }

        # Guardar reporte
        reporte_path = self.errors_dir / f"reporte_consolidacion_agentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(reporte_path, 'w', encoding='utf-8') as f:
            json.dump(reporte, f, indent=2, ensure_ascii=False)

        # Imprimir resumen
        logger.info("=" * 80)
        logger.info("REPORTE FINAL - CONSOLIDACIÓN DE AGENTES")
        logger.info("=" * 80)
        logger.info(f"Agentes procesados: {self.stats['total_agentes_raw']}")
        logger.info(f"Agentes únicos: {self.stats['agentes_unicos']}")
        logger.info(f"Duplicados eliminados: {self.stats['duplicados_eliminados']} ({reporte['metricas_calidad']['tasa_duplicacion']}%)")
        logger.info(f"Grupos de duplicados: {self.stats['grupos_duplicados']}")
        logger.info(f"Archivos procesados: {len(archivos_procesados)}")
        logger.info("")
        logger.info("Calidad de datos:")
        logger.info(f"  - Nombres completos: {calidad_datos['nombres_completos']} ({reporte['metricas_calidad']['porcentaje_nombres_completos']}%)")
        logger.info(f"  - Emails válidos: {calidad_datos['emails_validos']} ({reporte['metricas_calidad']['porcentaje_emails_validos']}%)")
        logger.info(f"  - Teléfonos válidos: {calidad_datos['telefonos_validos']} ({reporte['metricas_calidad']['porcentaje_telefonos_validos']}%)")
        logger.info(f"  - Empresas registradas: {calidad_datos['empresas_registradas']}")
        logger.info("")
        logger.info("Top 5 empresas:")
        for i, (empresa, count) in enumerate(top_empresas[:5], 1):
            logger.info(f"  {i}. {empresa}: {count} agentes")
        logger.info("")
        logger.info(f"Reporte guardado en: {reporte_path}")
        logger.info("=" * 80)

    def ejecutar_consolidacion(self):
        """
        Ejecutar el proceso completo de consolidación de agentes.
        """
        logger.info("Iniciando proceso de consolidación de agentes...")

        # 1. Procesar archivos y extraer agentes
        resultado_procesamiento = self.procesar_agentes_de_archivos()

        if not resultado_procesamiento['agentes']:
            logger.warning("No se encontraron agentes para procesar")
            return

        # 2. Deduplicar agentes
        agentes_unicos, grupos_duplicados = self.deduplicar_agentes(resultado_procesamiento['agentes'])

        # 3. Guardar archivo consolidado
        self.guardar_archivo_consolidado(
            agentes_unicos,
            grupos_duplicados,
            resultado_procesamiento['archivos_procesados']
        )

        # 4. Generar reporte final
        self.generar_reporte_final(
            agentes_unicos,
            resultado_procesamiento['archivos_procesados']
        )

        logger.info("Proceso de consolidación de agentes completado exitosamente")


def main():
    """Función principal para ejecutar el ETL."""
    # Configuración de directorios
    input_dir = "data/intermedio/procesados"
    output_dir = "data/intermedio/procesados"
    errors_dir = "data/intermedio/errores"

    # Crear instancia y ejecutar
    etl = AgentConsolidationETL(input_dir, output_dir, errors_dir)
    etl.ejecutar_consolidacion()


if __name__ == "__main__":
    main()